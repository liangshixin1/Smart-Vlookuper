# -*- coding: utf-8 -*-
"""
SMART-VLOOKUPER - Excel 字段匹配与 AI 自动化工具 (PyQt6)

主要功能：
- 模糊字段匹配与 COM 批量写入，自动保留原有单元格格式
- 内置 AI 助手：上传表格并描述需求后，实时预览流式生成的 Python 代码并在沙箱中执行
- 失败重试与可取消的进度提示，确保最终产出可正常打开的 Excel 文件

依赖：pip install pyqt6 pandas openpyxl pywin32 thefuzz
"""

import sys, os, re, warnings, json, subprocess, tempfile, threading, random, string, io, textwrap
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFileDialog, QMessageBox,
    QGridLayout, QGroupBox, QLabel, QLineEdit, QPushButton, QComboBox,
    QSpinBox, QHBoxLayout, QVBoxLayout, QListWidget, QListWidgetItem, QTableWidget,
    QTableWidgetItem, QAbstractItemView, QStyledItemDelegate, QRadioButton,
    QTabWidget, QDialog, QPlainTextEdit, QTextBrowser,
    QDialogButtonBox, QStackedWidget, QFormLayout, QSplitter
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QBrush, QColor, QAction, QTextCursor
from PyQt6.QtWebEngineWidgets import QWebEngineView
import gc
from thefuzz import fuzz
from copy import copy
from typing import List, Optional, Union, Mapping, Any

# —— 静默 openpyxl 的数据验证扩展警告 ——
warnings.filterwarnings(
    "ignore",
    message=".*Data Validation extension is not supported and will be removed.*",
    category=UserWarning,
    module="openpyxl"
)

# ===================== 全局设置管理 =====================

class AppSettings:
    """简单的JSON配置管理器，用于持久化用户偏好"""

    DEFAULTS = {
        "ai_api_key": "",
        "report_api_key": "",
        "theme": "dark",
        "engine_mode": "auto",
        "last_ai_export_path": ""
    }

    def __init__(self):
        self.config_dir = Path.home() / ".smart_vlookuper"
        self.path = self.config_dir / "settings.json"
        self.data = self.DEFAULTS.copy()
        self.load()

    def load(self):
        if self.path.exists():
            try:
                loaded = json.loads(self.path.read_text(encoding="utf-8"))
                if isinstance(loaded, dict):
                    self.data.update({k: loaded.get(k, v) for k, v in self.DEFAULTS.items()})
            except Exception:
                # 读取失败时保留默认配置
                pass

    def save(self):
        try:
            self.config_dir.mkdir(parents=True, exist_ok=True)
            self.path.write_text(json.dumps(self.data, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            # 配置写入失败不阻塞主流程
            pass

    def get(self, key, default=None):
        return self.data.get(key, default)

    def set(self, key, value):
        self.data[key] = value

    def update(self, **kwargs):
        self.data.update(kwargs)
        self.save()


PINK_ACCENT = "#F083A2"
PINK_ACCENT_HOVER = "#D96B8D"


EXPORT_FORMATS = {
    "excel": {
        "label": "Excel 工作簿 (.xlsx)",
        "extensions": [".xlsx", ".xlsm", ".xls"],
        "default_extension": ".xlsx",
        "dialog_filter": "Excel Files (*.xlsx *.xlsm *.xls)",
        "description": "Excel 工作簿"
    },
    "word": {
        "label": "Word 文档 (.docx)",
        "extensions": [".docx"],
        "default_extension": ".docx",
        "dialog_filter": "Word Documents (*.docx)",
        "description": "Word 文档"
    },
    "text": {
        "label": "文本文件 (.txt)",
        "extensions": [".txt", ".csv"],
        "default_extension": ".txt",
        "dialog_filter": "Text Files (*.txt *.csv)",
        "description": "文本文件"
    }
}


CODE_CHAR_NORMALIZATION = str.maketrans({
    "：": ":",
    "；": ";",
    "，": ",",
    "。": ".",
    "、": ",",
    "（": "(",
    "）": ")",
    "【": "[",
    "】": "]",
    "［": "[",
    "］": "]",
    "｛": "{",
    "｝": "}",
    "《": "<",
    "》": ">",
    "「": '"',
    "」": '"',
    "“": '"',
    "”": '"',
    "‘": "'",
    "’": "'",
    " ": " ",
    " ": " ",
    "　": " ",
})


def normalize_generated_code(code: str) -> str:
    """将常见的全角标点替换为Python可识别的半角符号"""
    if not code:
        return ""
    normalized = code.translate(CODE_CHAR_NORMALIZATION)
    normalized = normalized.replace("\r\n", "\n").replace("\ufeff", "")
    return normalized


PRIVACY_STATEMENT_TEXT = """最后更新日期：2025年9月26日

尊敬的用户：

感谢您使用本系统的AI助手功能。我们深知数据安全与隐私保护的重要性，尤其是您可能处理的个人敏感信息。我们遵循 “数据最小化”与 “隐私优先” 的核心原则。我们的AI助手不会将您的原始数据发送至任何第三方服务器进行处理。所有敏感信息的保护均在您的本地计算机上完成。当您上传Excel文件前，系统会在本地自动扫描表格的表头（列名称），并根据内置的规则库识别可能包含个人敏感信息的列。对于识别出的敏感列，系统会立即在您的电脑内存中**将该列的所有数据内容替换为计算机随机生成的、符合原始数据格式的虚假样本数据。**之后，系统将使用这份**已经过脱敏处理的、不包含任何真实个人信息的数据样本**来生成提供给AI模型的提示词。您原始的、真实的Excel文件数据在任何情况下都绝不会离开您的本地环境。所有生成的代码均在严格的隔离环境中运行，无法访问您的网络或计算机上的其他无关文件。

您有权知晓上述处理流程。使用AI功能即代表您同意此数据处理方式。如果您不希望进行任何形式的数据处理，请不要使用AI助手功能，软件的核心匹配功能仍可完全离线使用。在AI代码预览窗口中，您可以完整审查即将被执行的代码逻辑。

如果您对本声明或我们的数据安全有任何疑问，请访问我们的Github项目地址，并提交Issue"""


DARK_STYLESHEET = """
    QMainWindow { background: #0f172a; color: #f8fafc; }
    QMenuBar { background: #0f172a; color: #f8fafc; }
    QMenuBar::item { background: transparent; color: #f8fafc; padding: 4px 12px; }
    QMenuBar::item:selected { background: #1e293b; color: #f8fafc; }
    QMenu { background: #111827; color: #f8fafc; border: 1px solid #1f2937; }
    QMenu::item:selected { background: #1f2937; }
    QDialog, QMessageBox { background: #0f172a; color: #f8fafc; }
    QLabel { color: #f1f5f9; font-weight: 600; }
    QGroupBox { border: 1px solid #1f2937; border-radius: 10px; margin-top: 10px; padding: 10px; color: #f8fafc; font-weight: 600; }
    QGroupBox::title { subcontrol-origin: margin; subcontrol-position: top left; padding: 2px 6px; }
    QLineEdit, QComboBox, QSpinBox, QListWidget, QTableWidget { background: #111827; color: #f8fafc; border: 1px solid #374151; border-radius: 8px; padding: 6px; }
    QComboBox::drop-down { border: none; }
    QPushButton { background: %(accent)s; color: white; border: none; border-radius: 8px; padding: 8px 12px; font-weight: 600; }
    QPushButton:hover { background: %(accent_hover)s; }
    QPushButton:disabled { background: #334155; color: #9ca3af; }
    QRadioButton { color: #f8fafc; font-weight: normal; }
    QHeaderView::section { background: #0b1220; color: #f1f5f9; padding: 6px; border: none; }
    QTableWidget { gridline-color: #374151; }
    QTableWidget::item { padding-left: 5px; }
    QTabWidget::pane { border: 1px solid #1f2937; border-radius: 10px; }
    QTabBar::tab { background: #1e293b; color: #f1f5f9; padding: 6px 12px; margin: 2px; border-top-left-radius: 6px; border-top-right-radius: 6px; }
    QTabBar::tab:selected { background: %(accent)s; color: white; }
""" % {"accent": PINK_ACCENT, "accent_hover": PINK_ACCENT_HOVER}


LIGHT_STYLESHEET = """
    QMainWindow { background: #f8fafc; color: #0f172a; }
    QMenuBar { background: #f8fafc; color: #0f172a; }
    QMenuBar::item { background: transparent; color: #0f172a; padding: 4px 12px; }
    QMenuBar::item:selected { background: #e2e8f0; color: #0f172a; }
    QMenu { background: #ffffff; color: #0f172a; border: 1px solid #cbd5e1; }
    QMenu::item:selected { background: #e2e8f0; }
    QDialog, QMessageBox { background: #f8fafc; color: #0f172a; }
    QLabel { color: #0f172a; font-weight: 600; }
    QGroupBox { border: 1px solid #cbd5e1; border-radius: 10px; margin-top: 10px; padding: 10px; color: #0f172a; font-weight: 600; background: #ffffff; }
    QGroupBox::title { subcontrol-origin: margin; subcontrol-position: top left; padding: 2px 6px; }
    QLineEdit, QComboBox, QSpinBox, QListWidget, QTableWidget { background: #ffffff; color: #111827; border: 1px solid #cbd5e1; border-radius: 8px; padding: 6px; }
    QComboBox::drop-down { border: none; }
    QPushButton { background: %(accent)s; color: white; border: none; border-radius: 8px; padding: 8px 12px; font-weight: 600; }
    QPushButton:hover { background: %(accent_hover)s; }
    QPushButton:disabled { background: #cbd5e1; color: #94a3b8; }
    QRadioButton { color: #0f172a; font-weight: normal; }
    QHeaderView::section { background: #e2e8f0; color: #1e293b; padding: 6px; border: none; }
    QTableWidget { gridline-color: #cbd5e1; }
    QTableWidget::item { padding-left: 5px; }
    QTabWidget::pane { border: 1px solid #cbd5e1; border-radius: 10px; background: #ffffff; }
    QTabBar::tab { background: #e2e8f0; color: #1e293b; padding: 6px 12px; margin: 2px; border-top-left-radius: 6px; border-top-right-radius: 6px; }
    QTabBar::tab:selected { background: %(accent)s; color: white; }
""" % {"accent": PINK_ACCENT, "accent_hover": PINK_ACCENT_HOVER}

# ===================== 基础工具与模糊匹配 =====================

def norm_str(x):
    """规范化字符串，去除首尾空格"""
    return str(x).strip()

def dedup_columns(names):
    """为重复的列名添加后缀以保证唯一性"""
    seen, out = {}, []
    for n in names:
        base = n
        if base not in seen:
            seen[base] = 1
            out.append(base)
        else:
            seen[base] += 1
            out.append(f"{base} ({seen[base]})")
    return out


SENSITIVE_KEYWORDS = {
    "name": ["姓名", "名字", "name", "联系人", "客户姓名", "学生姓名"],
    "id": ["身份证", "身份证号", "身份证号码", "证件号", "证件号码", "id card", "idcard"],
    "phone": ["手机号", "手机", "电话", "联系电话", "phone", "mobile", "联系方式"],
    "address": ["地址", "住址", "家庭地址", "办公地址", "通讯地址", "联系地址", "address"]
}

FAKE_SURNAMES = list("赵钱孙李周吴郑王冯陈楮卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜戚谢邹喻柏水窦章云苏潘葛奚范彭郎鲁韦昌马苗凤花方俞任袁柳酆鲍史唐费廉岑薛雷贺倪汤滕殷罗毕郝邬安常乐于时傅皮卞齐康伍余元卜顾孟平黄和穆萧尹姚邵湛汪祁毛禹狄米贝明臧计成戴谈宋茅庞熊纪舒屈项祝董梁")
FAKE_GIVEN_SINGLE = ["明", "芳", "华", "敏", "伟", "磊", "娜", "强", "霞", "杰", "静", "勇", "婷", "超", "燕", "波", "丽", "凯", "睿", "佳"]
FAKE_GIVEN_DOUBLE = ["子涵", "梓萱", "浩然", "欣怡", "宇轩", "思远", "雅静", "晨曦", "博文", "诗涵", "靖雯", "曜文", "逸晨", "沐阳", "琪瑛", "思琪"]
FAKE_PROVINCES = ["京", "沪", "粤", "浙", "苏", "鲁", "川", "渝", "辽", "湘", "闽", "皖", "鄂", "津"]
FAKE_ADDRESS_SUFFIX = ["幸福路", "创新大道", "人民中路", "解放东街", "和平里", "阳光巷", "滨江道", "文化路", "建业街", "望江路", "星海路"]


def classify_sensitive_column(column_name: str):
    lowered = norm_str(column_name).lower()
    for category, keywords in SENSITIVE_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in lowered:
                return category
    return None


def _fake_chinese_name(length_hint: int = 2) -> str:
    surname = random.choice(FAKE_SURNAMES)
    if length_hint <= 2:
        given = random.choice(FAKE_GIVEN_SINGLE)
    else:
        given = random.choice(FAKE_GIVEN_DOUBLE)
    return surname + given


def _fake_id_like(text: str) -> str:
    stripped = re.sub(r"[^0-9Xx]", "", text)
    length = len(stripped) if stripped else max(len(text), 18)
    length = max(length, 6)
    body = "".join(random.choices(string.digits, k=max(length - 1, 1)))
    if length >= 18:
        return body[: length - 1] + random.choice(string.digits + "X")
    return body + random.choice(string.digits)


def _fake_phone_like(text: str) -> str:
    digits_only = re.sub(r"\D", "", text)
    length = len(digits_only) if digits_only else 11
    length = max(length, 8)
    if length >= 11:
        prefix = random.choice(["13", "15", "17", "18", "19"])
        remainder = "".join(random.choices(string.digits, k=length - len(prefix)))
        return (prefix + remainder)[:length]
    return "".join(random.choices(string.digits, k=length))


def _fake_address() -> str:
    return (
        f"{random.choice(FAKE_PROVINCES)}省"
        f"{random.choice(['城区', '新区', '高新区', '经济开发区', '中心区'])}"
        f"{random.choice(FAKE_ADDRESS_SUFFIX)}{random.randint(1, 299)}号"
    )


def _fake_generic(text: str) -> str:
    length = len(text)
    if length <= 0:
        return ""
    choices = string.ascii_letters + string.digits
    return "".join(random.choices(choices, k=length))


def generate_fake_value(value, category: str) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if category == "name":
        return _fake_chinese_name(len(text))
    if category == "id":
        return _fake_id_like(text)
    if category == "phone":
        return _fake_phone_like(text)
    if category == "address":
        return _fake_address()
    return _fake_generic(text)


def desensitize_dataframe(df: pd.DataFrame):
    sanitized = df.copy()
    sensitive_columns = {}
    for col in sanitized.columns:
        sanitized[col] = sanitized[col].apply(lambda x: "" if pd.isna(x) else str(x))
        category = classify_sensitive_column(col)
        if category:
            sanitized[col] = sanitized[col].apply(lambda v: generate_fake_value(v, category))
            sensitive_columns[col] = category
    return sanitized, sensitive_columns


def generate_dataframe_fingerprint(
    df: pd.DataFrame,
    max_value_columns: int = 6,
    max_category_items: int = 6,
    max_chars: int = 4000
) -> str:
    """构建数据指纹，帮助LLM了解完整数据集的结构与分布"""

    if df is None or not isinstance(df, pd.DataFrame):
        return "(无可用指纹信息)"

    sections: list[str] = []

    buffer = io.StringIO()
    try:
        df.info(buf=buffer)
    except Exception:
        info_text = "无法生成 df.info() 摘要。"
    else:
        info_text = buffer.getvalue().strip()
    if info_text:
        sections.append("[DataFrame.info]\n" + info_text)

    try:
        numeric_cols = df.select_dtypes(include=["number"])
        if not numeric_cols.empty:
            desc = numeric_cols.describe().transpose().head(max_value_columns)
            sections.append("[数值列统计]\n" + desc.to_string())
    except Exception:
        pass

    try:
        datetime_cols = df.select_dtypes(include=["datetime64[ns]", "datetime64[ns, tz]"])
        if not datetime_cols.empty:
            lines = []
            for col in list(datetime_cols.columns)[:max_value_columns]:
                series = datetime_cols[col].dropna()
                if series.empty:
                    lines.append(f"{col}: 无有效日期")
                else:
                    lines.append(
                        f"{col}: {series.min()} ~ {series.max()} (样本 {series.size})"
                    )
            if lines:
                sections.append("[日期列范围]\n" + "\n".join(lines))
    except Exception:
        pass

    try:
        category_cols = df.select_dtypes(include=["object", "category", "bool"])
        if not category_cols.empty:
            cat_lines = []
            for col in list(category_cols.columns)[:max_value_columns]:
                try:
                    vc = category_cols[col].astype(str).value_counts(dropna=False).head(max_category_items)
                except Exception:
                    continue
                if vc.empty:
                    cat_lines.append(f"{col}: （无数据）")
                    continue
                formatted = []
                for idx, cnt in vc.items():
                    value = str(idx)
                    if len(value) > 40:
                        value = value[:40] + "…"
                    formatted.append(f"{value}({cnt})")
                cat_lines.append(f"{col}: " + ", ".join(formatted))
            if cat_lines:
                sections.append("[类别列分布]\n" + "\n".join(cat_lines))
    except Exception:
        pass

    fingerprint = "\n\n".join(sections).strip()
    if not fingerprint:
        fingerprint = "(无可用指纹信息)"
    if len(fingerprint) > max_chars:
        fingerprint = fingerprint[:max_chars] + "\n…(指纹信息已截断)"
    return fingerprint

def find_best_match(target_field, source_fields, threshold=85):
    """
    【新增】使用多层策略为目标字段查找最佳的源字段匹配
    """
    norm_tgt = target_field.strip().lower()
    
    # 预定义的近义词字典
    synonyms = {
        "手机号码": ["手机号", "联系方式", "电话", "联系电话"],
        "毕业院校": ["毕业学校", "学校", "院校"],
        "身份证": ["身份证号", "身份证号码", "id", "证件号码"],
        "工号": ["员工编号", "员工id", "eid"],
        "姓名": ["员工姓名"],
        "学历": ["最高学历"]
    }
    
    best_match = None
    highest_score = 0

    # 策略1: 精确与标准化精确匹配
    for src in source_fields:
        if src == target_field: return src
        if norm_str(src).lower() == norm_tgt:
            best_match = src
            highest_score = 101  # 赋予最高优先级

    if best_match: return best_match

    # 策略2: 近义词匹配
    synonym_group = []
    for key, values in synonyms.items():
        if norm_tgt == key.lower() or norm_tgt in [v.lower() for v in values]:
            synonym_group.extend([key.lower()] + [v.lower() for v in values])
            break
    
    if synonym_group:
        for src in source_fields:
            norm_src = norm_str(src).lower()
            if norm_src in synonym_group:
                score = fuzz.ratio(norm_tgt, norm_src)
                if score > highest_score:
                    highest_score = score
                    best_match = src
    
    if best_match: return best_match

    # 策略3: 模糊相似度匹配
    for src in source_fields:
        score = fuzz.partial_ratio(norm_tgt, norm_str(src).lower())
        if score > highest_score:
            highest_score = score
            best_match = src
    
    if highest_score >= threshold:
        return best_match
    
    return None

def read_excel_dataframe(path: Path, sheet_name: str, header_row: int, data_start_col: int, drop_all_blank_rows: bool=True):
    """
    【关键】安全的Excel文件读取，通过 dtype=str 保证数据格式不变
    """
    try:
        # dtype=str 是保证'00001'和手机号不被转换的关键
        df = pd.read_excel(
            path, 
            sheet_name=sheet_name, 
            header=header_row - 1,
            engine="openpyxl", 
            dtype=str
        ).fillna('')
        
        if data_start_col > 1:
            df = df.iloc[:, data_start_col - 1:]

        if drop_all_blank_rows: 
            df = df.dropna(how="all")
        
        raw_cols = [norm_str(c) or f"(无名列){get_column_letter(i + data_start_col)}" for i, c in enumerate(df.columns)]
        df.columns = dedup_columns(raw_cols)
        
        gc.collect()
        return df
        
    except Exception as e:
        gc.collect()
        raise e

def get_sheet_names_safely(path: Path):
    """安全获取工作表名称，确保文件句柄正确关闭"""
    wb = None
    try:
        wb = load_workbook(path, read_only=True, keep_vba=True)
        return wb.sheetnames.copy()
    finally:
        if wb: wb.close()
        gc.collect()

def suggest_index_choice(columns):
    """根据常见词推荐索引列"""
    prefer = {"姓名", "name", "Name", "Full Name", "姓名（必填）", "工号"}
    for w in prefer:
        for c in columns:
            if norm_str(c) == w: return c
    return columns[0] if columns else None

def auto_detect_header_start(path: Path, sheet_name: str, max_rows: int = 50):
    """自动识别表头行和数据起始列"""
    try:
        df = pd.read_excel(
            path,
            sheet_name=sheet_name,
            header=None,
            nrows=max_rows,
            engine="openpyxl",
            dtype=str
        ).fillna('')
    except Exception:
        return 1, 1

    keywords = ["姓名", "名称", "工号", "号码", "电话", "时间", "地址", "金额"]
    best_row, best_score = 0, -1

    for idx, row in df.iterrows():
        cells = [norm_str(c) for c in row.tolist()]
        non_empty = [c for c in cells if c]
        if not non_empty:
            continue
        non_empty_count = len(non_empty)
        text_cells = [c for c in non_empty if not re.fullmatch(r"-?\d+(?:\.\d+)?", c)]
        text_ratio = len(text_cells) / non_empty_count
        keyword_hits = sum(1 for c in non_empty for kw in keywords if kw in c)
        score = non_empty_count + text_ratio * 5 + keyword_hits * 10
        if score > best_score:
            best_score = score
            best_row = idx

    header_row = best_row + 1
    header_cells = [norm_str(c) for c in df.iloc[best_row].tolist()]
    start_col = 1
    for j, val in enumerate(header_cells):
        if val:
            start_col = j + 1
            break

    return header_row, start_col

class ComboDelegate(QStyledItemDelegate):
    """用于表格内嵌下拉框的代理"""
    def __init__(self, parent, options):
        super().__init__(parent)
        self.options = options
    def createEditor(self, parent, option, index):
        combo = QComboBox(parent)
        combo.addItems(self.options)
        return combo
    def setEditorData(self, editor, index):
        value = index.model().data(index, Qt.ItemDataRole.EditRole)
        i = editor.findText(value) if value else -1
        editor.setCurrentIndex(i if i >= 0 else 0)
    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText(), Qt.ItemDataRole.EditRole)

# ===================== Excel 写入与导出 =====================

def excel_com_write_and_save_optimized(tgt_path: Path, tgt_sheet: str, out_path: Path,
                                     df_src: pd.DataFrame, df_tgt: pd.DataFrame, src_map: pd.DataFrame,
                                     mapping: list, tgt_field_to_col: dict, tgt_data_start_row: int,
                                     overwrite_all: bool):
    """【性能核心】使用批量写入提升COM性能，按列批量操作而非逐格"""
    try: 
        import win32com.client as win32
    except Exception as e: 
        raise RuntimeError("未安装或无法加载 pywin32。") from e
    
    excel = None
    wb = None
    total_found, total_write = 0, 0
    
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.Calculation = win32.constants.xlCalculationManual
        
        wb = excel.Workbooks.Open(str(tgt_path))
        ws = wb.Worksheets(tgt_sheet)
        
        idx_to_row = {row.get("_IDX_", ""): tgt_data_start_row + row_offset
                      for row_offset, (_, row) in enumerate(df_tgt.iterrows())
                      if row.get("_IDX_", "") in src_map.index}
        total_found = len(idx_to_row)
        
        for tgt_field, src_field in mapping:
            if not src_field or src_field not in df_src.columns: continue
            tgt_col = tgt_field_to_col.get(tgt_field)
            if not tgt_col: continue
            
            updates = []
            for idx_val, excel_row in idx_to_row.items():
                val = src_map.loc[idx_val].get(src_field)
                if val == '' or pd.isna(val): continue
                
                if not overwrite_all:
                    cell = ws.Cells(excel_row, tgt_col)
                    if cell.Value is not None and str(cell.Value).strip() != "": continue
                
                updates.append((excel_row, val))
            
            if updates:
                updates.sort(key=lambda x: x[0])
                i = 0
                while i < len(updates):
                    start_row, values = updates[i][0], [updates[i][1]]
                    j = i + 1
                    while j < len(updates) and updates[j][0] == updates[j-1][0] + 1:
                        values.append(updates[j][1]); j += 1
                    
                    if len(values) == 1:
                        ws.Cells(start_row, tgt_col).Value = values[0]
                    else:
                        end_row = start_row + len(values) - 1
                        range_addr = f"{get_column_letter(tgt_col)}{start_row}:{get_column_letter(tgt_col)}{end_row}"
                        ws.Range(range_addr).Value = [[v] for v in values]
                    
                    total_write += len(values)
                    i = j
        
        excel.Calculation = win32.constants.xlCalculationAutomatic
        excel.ScreenUpdating = True
        
        ext = out_path.suffix.lower()
        ff = 52 if ext == ".xlsm" else 51 if ext == ".xlsx" else None
        wb.SaveAs(str(out_path), FileFormat=ff) if ff else wb.SaveAs(str(out_path))
        
    finally:
        if wb:
            try: wb.Close(SaveChanges=False)
            except: pass
        if excel:
            try: excel.Quit()
            except: pass
        gc.collect()
    
    return total_found, total_write

# ===================== AI 助手 =====================

def summarize_error(msg: str, columns=None) -> str:
    """提取错误信息的关键信息，减少token占用"""
    if not msg:
        return "未知错误"
    line = msg.strip().splitlines()[-1]
    m = re.search(r"KeyError: '([^']+)'", line)
    if m:
        info = f"执行因 KeyError 失败，缺少列 '{m.group(1)}'"
        if columns:
            info += f"。可用列: {list(columns)}"
        return info
    return line


class AIWorker(QThread):
    progress = pyqtSignal(str)
    success = pyqtSignal(str)
    error = pyqtSignal(str)
    code_stream = pyqtSignal(str)
    code_ready = pyqtSignal(str)

    def __init__(self, api_key, tables, instruction, temperature, language, output_path, output_format, conversation_history=None):
        super().__init__()
        self.api_key = api_key
        self.tables = [str(p) for p in tables]
        self.instruction = instruction
        self.temperature = temperature
        self.language = language
        self.output_path = Path(output_path)
        self.output_format = output_format if output_format in EXPORT_FORMATS else "excel"
        self.approval_event = threading.Event()
        self.history = conversation_history or []

    def approve_execution(self):
        self.approval_event.set()

    def _finalize_success(self, all_columns):
        expected_path = self.output_path
        if not expected_path.exists():
            return False, f"未生成指定路径的文件：{expected_path}"

        try:
            if self.output_format == "excel":
                suffix = expected_path.suffix.lower()
                if suffix in {".xlsx", ".xlsm"}:
                    load_workbook(expected_path).close()
                elif suffix == ".xls":
                    if expected_path.stat().st_size <= 0:
                        return False, "生成的Excel文件内容为空。"
                else:
                    return False, f"文件扩展名与期望的Excel格式不符：{suffix}"
            else:
                if expected_path.stat().st_size <= 0:
                    return False, "生成的文件内容为空。"
        except Exception as e:
            return False, summarize_error(str(e), all_columns)

        self.success.emit(str(expected_path))
        return True, None

    def run(self):
        try:
            from openai import OpenAI
        except Exception as e:
            self.error.emit(f"未安装openai库: {e}")
            return

        self.progress.emit("读取表格示例...")
        table_texts = []
        all_columns = set()
        for p in self.tables:
            if self.isInterruptionRequested():
                self.error.emit("已取消")
                return
            try:
                df = pd.read_excel(p, dtype=str).fillna("")
            except Exception as e:
                self.error.emit(f"无法读取表格：{p} - {e}")
                return
            sanitized_df, sensitive_map = desensitize_dataframe(df)
            sample = sanitized_df.head(5).to_csv(sep='\t', index=False).strip()
            cols = ", ".join(map(str, df.columns))
            all_columns.update(df.columns)
            fingerprint = ""
            try:
                df_full = pd.read_excel(p)
            except Exception:
                df_full = None
            fingerprint = generate_dataframe_fingerprint(df_full if df_full is not None else df)
            block_lines = [
                f"## {Path(p).name}",
                f"路径: {p}",
                f"列: {cols}",
            ]
            if sensitive_map:
                block_lines.append(f"已自动脱敏列: {', '.join(sensitive_map.keys())}")
            block_lines.append("示例:")
            block_lines.append(sample or "(空表)")
            if fingerprint:
                block_lines.append("数据指纹:")
                block_lines.append(fingerprint)
            table_texts.append("\n".join(block_lines))

        try:
            self.output_path.parent.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            self.error.emit(f"无法创建导出目录：{e}")
            return

        output_path_str = str(self.output_path)
        tables_json = json.dumps(self.tables, ensure_ascii=False) if self.tables else "[]"
        table_list_string = "\n".join(self.tables)
        table_info_text = "\n\n".join(table_texts) if table_texts else "（未选择表格）"
        language_key = self.language.lower()

        format_info = EXPORT_FORMATS.get(self.output_format, EXPORT_FORMATS["excel"])
        format_desc = format_info["description"]
        extensions_text = ", ".join(format_info["extensions"])

        history_lines = []
        for msg in self.history:
            role = msg.get("role")
            content = (msg.get("content") or "").strip()
            if not content:
                continue
            prefix = "用户" if role == "user" else "助手"
            history_lines.append(f"{prefix}：{content}")

        conversation_text = "\n".join(history_lines)
        if conversation_text:
            conversation_block = (
                "以下是之前的对话历史，请在继续编写代码时保持上下文一致：\n"
                f"{conversation_text}\n\n"
            )
        else:
            conversation_block = ""

        task_block = (
            f"{conversation_block}"
            f"当前用户最新指令：\n{self.instruction}\n\n"
            f"可用的Excel表格信息：\n{table_info_text}\n\n"
            f"目标输出文件类型：{format_desc}（可用扩展名：{extensions_text}）。\n"
            "所有示例数据均已在本地完成脱敏处理，仅用于展示列结构。\n"
            "以上绝对路径已由系统自动注入运行环境，无需等待用户重复提供。请直接使用这些路径，避免出现“Required Files Not Found”错误。"
        )

        if language_key == "vba":
            env_instructions = (
                "你将获得若干Excel文件的路径、列名以及前5行数据示例。请仅生成VBA代码。\n"
                "生成代码时必须直接使用提供的完整文件路径，避免因路径错误导致“Required Files Not Found”。\n"
                "必须声明一个入口宏：Sub ProcessTables(tableList As String, outputPath As String)。\n"
                "参数 tableList 为使用换行分隔的完整Excel路径字符串；outputPath 为结果Excel文件的完整保存路径。\n"
                f"运行环境会调用 ProcessTables(tableList, outputPath)，并且 outputPath 始终为：{output_path_str}（需生成 {format_desc}，扩展名：{extensions_text}）。\n"
                "请在宏内拆分 tableList，按需打开并处理这些工作簿，最终将结果保存到 outputPath 指定的路径。\n"
                "不要弹出对话框或依赖任何交互，也不要修改除结果文件外的其他文件。\n"
                "交互要求：在构思代码时，请同步准备一段代码执行成功后的反馈话术，先确认任务完成，再礼貌询问用户下一步需求。该话术无需写入代码，将在我们系统中展示给用户。\n"
                "仅返回纯VBA代码，不要包含```标记或额外说明。"
            )
            retry_suffix = "请仅返回修正后的VBA代码。"
        else:
            env_instructions = (
                "你将获得若干Excel文件的路径、列名以及前5行数据示例。请仅生成可直接运行的Python代码以满足用户需求。\n"
                "生成代码时必须直接使用提供的完整文件路径，避免因路径错误导致“Required Files Not Found”。\n"
                "运行环境提供了三个环境变量：AI_TABLE_PATHS（JSON数组，包含所有Excel完整路径）、AI_OUTPUT_PATH（结果文件完整路径）与 AI_OUTPUT_FORMAT（目标文件格式关键字）。\n"
                f"输出文件的目标路径固定为：{output_path_str}，并且期望生成 {format_desc}（扩展名：{extensions_text}）。请务必按此格式保存。\n"
                "代码完成后必须打印单行JSON，例如 print(json.dumps({'status':'success','output_path': output_path}, ensure_ascii=False))。\n"
                "不要输出任何解释或额外文本，也不要包含```代码块标记。\n"
                "交互要求：在规划代码时，请准备一段成功完成任务后的反馈话术，先确认任务完成，再询问用户下一步需求。该话术无需写入代码，将由我们的系统在执行成功后提示用户。"
            )
            retry_suffix = "请仅返回修正后的Python代码。"

        base_prompt = env_instructions + "\n\n" + task_block


        client = OpenAI(api_key=self.api_key, base_url="https://api.deepseek.com/v1")
        attempt, last_err, last_code = 0, None, ""
        while attempt < 3:
            if self.isInterruptionRequested():
                self.error.emit("已取消")
                return

            self.approval_event.clear()
            prompt = base_prompt if not last_err else base_prompt + f"\n\n上次执行错误：{last_err}\n{retry_suffix}"
            self.progress.emit(f"调用模型生成{self.language}代码...尝试{attempt + 1}/3")
            try:
                response = client.chat.completions.create(
                    model="deepseek-chat",
                    temperature=self.temperature,
                    messages=[
                        {"role": "system", "content": "You are a helpful assistant"},
                        {"role": "user", "content": prompt}
                    ],
                    stream=True
                )
            except Exception as e:
                self.error.emit(str(e))
                return

            code = ""
            try:
                for chunk in response:
                    if self.isInterruptionRequested():
                        self.error.emit("已取消")
                        return
                    # Each streamed token arrives as a ChoiceDelta object; get its text safely
                    delta_obj = chunk.choices[0].delta
                    delta = getattr(delta_obj, "content", None)

                    if delta:
                        code += delta
                        self.code_stream.emit(code)
            except Exception as e:
                self.error.emit(str(e))
                return

            code = code.strip()
            if code.startswith("```"):
                code = "\n".join(code.splitlines()[1:-1])
            last_code = code

            self.code_ready.emit(code)
            self.progress.emit("等待用户确认...")
            self.approval_event.wait()
            if self.isInterruptionRequested():
                self.error.emit("已取消")
                return

            if self.output_path.exists():
                try:
                    self.output_path.unlink()
                except Exception as e:
                    last_err = f"无法覆盖现有输出文件：{e}"
                    attempt += 1
                    continue

            self.progress.emit("执行代码...")
            if language_key == "vba":
                try:
                    execute_vba_module(code, table_list_string, self.output_path)
                except Exception as e:
                    last_err = summarize_error(str(e), all_columns)
                    attempt += 1
                    continue

                expected_path = self.output_path
                if expected_path.exists():
                    ok, err_text = self._finalize_success(all_columns)
                    if ok:
                        return
                    last_err = err_text
                else:
                    last_err = f"未生成指定路径的文件：{expected_path}"
            else:
                with tempfile.TemporaryDirectory() as td:
                    script = Path(td) / "script.py"
                    script.write_text(code, encoding="utf-8")
                    env = os.environ.copy()
                    env.setdefault("PYTHONPATH", "")
                    env["AI_TABLE_PATHS"] = tables_json
                    env["AI_TABLE_LIST"] = table_list_string
                    env["AI_OUTPUT_PATH"] = output_path_str
                    env["AI_OUTPUT_FORMAT"] = self.output_format
                    env["AI_OUTPUT_EXTENSIONS"] = json.dumps(format_info["extensions"], ensure_ascii=False)
                    env["AI_INSTRUCTION_TEXT"] = self.instruction
                    try:
                        proc = subprocess.run(
                            [sys.executable, str(script)],
                            capture_output=True,
                            text=True,
                            cwd=td,
                            env=env,
                        )
                    except Exception as e:
                        last_err = str(e)
                        attempt += 1
                        continue

                stdout = proc.stdout.strip()
                stderr = proc.stderr.strip()
                if proc.returncode != 0:
                    last_err = summarize_error(stderr or stdout, all_columns)
                else:
                    expected_path = self.output_path
                    if expected_path.exists():
                        ok, err_text = self._finalize_success(all_columns)
                        if ok:
                            return
                        last_err = err_text
                    else:
                        result_json = None
                        if stdout:
                            try:
                                result_json = json.loads(stdout.splitlines()[-1])
                            except Exception:
                                result_json = None
                        if result_json and result_json.get("output_path"):
                            candidate = Path(result_json.get("output_path"))
                            if candidate.exists():
                                last_err = f"模型在 {candidate} 生成了文件，请将结果保存至指定路径：{expected_path}"
                            else:
                                last_err = f"未能在指定路径生成结果文件：{expected_path}"
                        else:
                            last_err = summarize_error(stdout or stderr, all_columns)


            attempt += 1

        self.error.emit((last_err or "执行失败") + f"\n\n最后的代码:\n{last_code}")


class IntentWorker(QThread):
    """识别上传表格的常见处理意图"""

    results = pyqtSignal(list)
    error = pyqtSignal(str)

    def __init__(self, api_key: str, tables):
        super().__init__()
        self.api_key = api_key
        self.tables = list(tables) if tables is not None else []

    def _read_dataframe(self, path: Path, as_string: bool) -> pd.DataFrame:
        if as_string:
            if path.suffix.lower() == ".csv":
                df = pd.read_csv(path, dtype=str)
            else:
                df = pd.read_excel(path, dtype=str, engine="openpyxl")
            return df.fillna("")
        if path.suffix.lower() == ".csv":
            return pd.read_csv(path)
        return pd.read_excel(path, engine="openpyxl")

    def _prepare_entry(self, item) -> dict:
        path_obj: Optional[Path] = None
        dataframe_input: Optional[pd.DataFrame] = None
        display_name = ""

        if isinstance(item, Mapping):
            raw_path = item.get("path")
            if raw_path:
                path_obj = Path(raw_path)
            df_candidate = item.get("dataframe")
            if isinstance(df_candidate, pd.DataFrame):
                dataframe_input = df_candidate
            elif df_candidate is not None:
                raise TypeError("dataframe 字段必须是 pandas.DataFrame")
            display_name = str(item.get("name") or item.get("display_name") or item.get("label") or "")
        elif isinstance(item, pd.DataFrame):
            dataframe_input = item
        elif isinstance(item, (str, Path)):
            path_obj = Path(item)
        else:
            raise TypeError("不支持的表格类型")

        if not display_name:
            if path_obj is not None:
                display_name = path_obj.name
            else:
                display_name = "数据集"

        if dataframe_input is not None:
            df_full = dataframe_input.copy()
            df_preview = dataframe_input.fillna("").astype(str)
        else:
            if path_obj is None:
                raise ValueError("缺少表格路径或数据。")
            try:
                df_preview = self._read_dataframe(path_obj, as_string=True)
            except Exception as e:
                raise RuntimeError(f"无法读取表格：{path_obj} - {e}") from e
            try:
                df_full = self._read_dataframe(path_obj, as_string=False)
            except Exception:
                df_full = df_preview.copy()

        return {
            "name": display_name,
            "path": path_obj,
            "df_preview": df_preview,
            "df_full": df_full,
        }

    def run(self):
        if not self.tables:
            self.results.emit([])
            return

        try:
            from openai import OpenAI
        except Exception as e:
            self.error.emit(f"未安装openai库: {e}")
            return

        table_chunks = []
        for item in self.tables:
            if self.isInterruptionRequested():
                return
            try:
                entry = self._prepare_entry(item)
            except Exception as e:
                self.error.emit(str(e))
                return

            df_preview = entry["df_preview"]
            sanitized_df, sensitive_map = desensitize_dataframe(df_preview)
            sample_csv = sanitized_df.head(5).to_csv(index=False).strip()
            columns = json.dumps([str(c) for c in df_preview.columns], ensure_ascii=False)
            df_full = entry.get("df_full")
            fingerprint = generate_dataframe_fingerprint(df_full if df_full is not None else df_preview)
            chunk_lines = [
                f"数据集: {entry['name']}",
                f"列名: {columns}",
            ]
            if sensitive_map:
                chunk_lines.append(f"已自动脱敏列: {', '.join(sensitive_map.keys())}")
            chunk_lines.append("数据样本:")
            chunk_lines.append(sample_csv or "(空表)")
            if fingerprint:
                chunk_lines.append("数据指纹:")
                chunk_lines.append(fingerprint)
            table_chunks.append("\n".join(chunk_lines))

        prompt_text = "以下是一个或多个表格的结构与数据示例：\n\n" + "\n\n".join(table_chunks)

        client = OpenAI(api_key=self.api_key, base_url="https://api.deepseek.com/v1")
        try:
            response = client.chat.completions.create(
                model="deepseek-chat",
                temperature=0.0,
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "你是一个精通Excel和数据处理的专家助理。你的任务是根据用户上传的表格的列名和少量示例数据，预测用户可能想要执行的几个最常见的操作。"
                            "请仅以JSON数组的格式返回5个操作意图，每个意图为2-5个字且动词开头。"
                        ),
                    },
                    {"role": "user", "content": prompt_text},
                ],
            )
        except Exception as e:
            if not self.isInterruptionRequested():
                self.error.emit(str(e))
            return

        if self.isInterruptionRequested():
            return

        content = ""
        try:
            content = (response.choices[0].message.content or "").strip()
        except Exception:
            self.error.emit("意图识别返回内容为空")
            return

        intents = None
        try:
            intents = json.loads(content)
        except Exception:
            try:
                start = content.find("[")
                end = content.rfind("]")
                if start != -1 and end != -1 and end > start:
                    intents = json.loads(content[start : end + 1])
            except Exception:
                intents = None

        if not isinstance(intents, list):
            self.error.emit("意图识别结果解析失败")
            return

        cleaned = []
        for item in intents:
            text = str(item).strip()
            if text:
                cleaned.append(text)

        self.results.emit(cleaned[:5])


def execute_vba_module(code: str, table_payload: str, output_path: Path):
    """在临时工作簿中插入并执行VBA代码"""
    try:
        import win32com.client as win32
    except Exception as e:
        raise RuntimeError("未安装或无法加载 pywin32。") from e

    excel = None
    wb = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Add()
        except Exception as e:
            raise RuntimeError(f"无法创建临时工作簿：{e}") from e

        try:
            module = wb.VBProject.VBComponents.Add(1)
        except Exception as e:
            raise RuntimeError("无法访问Excel VBA项目，请在Excel选项中启用“信任对VBA项目对象模型的访问”。") from e

        module.CodeModule.AddFromString(code)
        try:
            excel.Run("ProcessTables", table_payload, str(output_path))
        except Exception as e:
            raise RuntimeError(f"VBA 执行失败：{e}") from e

    finally:
        if wb:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        gc.collect()


class AIAssistantWidget(QWidget):
    def __init__(self, settings: AppSettings, parent=None):
        super().__init__(parent)
        self.settings = settings
        self.setMinimumSize(800, 600)
        self.tables = []
        self.conversation_messages = []
        self.worker = None
        self.awaiting_execution = False
        self._last_status_text = ""
        self.intent_worker = None
        self.recommend_buttons = []
        self.recommend_placeholder = None

        main_layout = QHBoxLayout(self)
        main_layout.setSpacing(12)

        history_group = QGroupBox("历史对话")
        history_group.setMinimumWidth(220)
        history_layout = QVBoxLayout(history_group)
        self.history_list = QListWidget()
        self.history_list.setAlternatingRowColors(True)
        self.history_list.setWordWrap(True)
        self.history_list.itemDoubleClicked.connect(self.show_history_detail)
        history_layout.addWidget(self.history_list)
        history_btn_row = QHBoxLayout()
        self.btn_new_session = QPushButton("开启新对话")
        self.btn_new_session.clicked.connect(self.start_new_session)
        history_btn_row.addWidget(self.btn_new_session)
        self.btn_clear_history = QPushButton("清空历史")
        self.btn_clear_history.clicked.connect(self.clear_history)
        history_btn_row.addWidget(self.btn_clear_history)
        history_layout.addLayout(history_btn_row)
        main_layout.addWidget(history_group, 1)

        center_group = QGroupBox("对话配置")
        center_layout = QVBoxLayout(center_group)
        center_layout.addWidget(QLabel("使用场景:"))
        self.scenario_combo = QComboBox()
        self.scenario_combo.addItems([
            "代码生成/数学解题",
            "数据抽取/分析",
            "通用对话",
            "翻译",
            "创意类写作/诗歌创作"
        ])
        center_layout.addWidget(self.scenario_combo)

        self.btn_add_table = QPushButton("添加表格")
        self.btn_add_table.clicked.connect(self.add_table)
        center_layout.addWidget(self.btn_add_table)

        self.table_list = QListWidget()
        self.table_list.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.table_list.setMinimumHeight(120)
        center_layout.addWidget(self.table_list, 1)

        center_layout.addWidget(QLabel("生成代码语言:"))
        self.language_combo = QComboBox()
        self.language_combo.addItems(["Python", "VBA"])
        center_layout.addWidget(self.language_combo)

        center_layout.addWidget(QLabel("导出文件格式:"))
        self.export_format_combo = QComboBox()
        for key, info in EXPORT_FORMATS.items():
            self.export_format_combo.addItem(info["label"], userData=key)
        self.export_format_combo.currentIndexChanged.connect(self.on_export_format_changed)
        center_layout.addWidget(self.export_format_combo)

        center_layout.addWidget(QLabel("导出结果路径:"))
        path_layout = QHBoxLayout()
        self.output_edit = QLineEdit()
        self.output_edit.setPlaceholderText("请选择AI生成结果的保存路径")
        default_output = self.settings.get("last_ai_export_path", "") or ""
        if default_output:
            self.output_edit.setText(default_output)
        self._sync_format_combo_with_path(default_output)
        btn_browse = QPushButton("浏览…")
        btn_browse.clicked.connect(self.browse_output)
        path_layout.addWidget(self.output_edit, 1)
        path_layout.addWidget(btn_browse)
        center_layout.addLayout(path_layout)

        center_layout.addWidget(QLabel("对话输入:"))
        self.message_edit = QPlainTextEdit()
        self.message_edit.setPlaceholderText("请用自然语言描述下一步操作，例如：合并客户表的电话字段。系统会自动提供所选表格的完整路径。")
        self.message_edit.setMinimumHeight(140)
        center_layout.addWidget(self.message_edit)

        self.privacy_label = QLabel("<a href='#'>用户隐私声明</a>")
        self.privacy_label.setStyleSheet("color: #3b82f6;")
        self.privacy_label.setTextFormat(Qt.TextFormat.RichText)
        self.privacy_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
        self.privacy_label.setCursor(Qt.CursorShape.PointingHandCursor)
        self.privacy_label.linkActivated.connect(self.show_privacy_statement)
        center_layout.addWidget(self.privacy_label, alignment=Qt.AlignmentFlag.AlignLeft)

        self.recommend_container = QWidget()
        self.recommend_layout = QHBoxLayout(self.recommend_container)
        self.recommend_layout.setContentsMargins(0, 0, 0, 0)
        self.recommend_layout.setSpacing(6)
        center_layout.addWidget(self.recommend_container)
        self._reset_recommend_placeholder("推荐操作将在此显示")

        self.btn_send = QPushButton("发送指令")
        self.btn_send.clicked.connect(self.send_message)
        center_layout.addWidget(self.btn_send)
        center_layout.addStretch()
        main_layout.addWidget(center_group, 2)

        preview_group = QGroupBox("预览")
        preview_layout = QVBoxLayout(preview_group)
        self.preview_tabs = QTabWidget()
        self.code_preview = QPlainTextEdit()
        self.code_preview.setReadOnly(True)
        self.code_preview.setPlaceholderText("AI生成的代码将实时显示在此处。")
        self.preview_tabs.addTab(self.code_preview, "代码预览")
        self.log_view = QPlainTextEdit()
        self.log_view.setReadOnly(True)
        self.log_view.setPlaceholderText("执行日志与提示将显示在此处。")
        self.preview_tabs.addTab(self.log_view, "执行日志")
        self.table_preview = QTableWidget()
        self.table_preview.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_preview.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_preview.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.table_preview.horizontalHeader().setStretchLastSection(True)
        self.preview_tabs.addTab(self.table_preview, "表格预览")
        preview_layout.addWidget(self.preview_tabs, 1)
        self.status_label = QLabel("等待指令…")
        preview_layout.addWidget(self.status_label)
        btn_row = QHBoxLayout()
        self.btn_execute = QPushButton("执行生成代码")
        self.btn_execute.setEnabled(False)
        self.btn_execute.clicked.connect(self.exec_generated_code)
        self.btn_cancel = QPushButton("取消当前操作")
        self.btn_cancel.clicked.connect(self.cancel_current)
        btn_row.addWidget(self.btn_execute)
        btn_row.addWidget(self.btn_cancel)
        preview_layout.addLayout(btn_row)
        main_layout.addWidget(preview_group, 2)

    def add_table(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "选择表格", "", "Excel Files (*.xlsx *.xlsm *.xls)")
        if not paths:
            return
        new_paths = []
        for p in paths:
            if p not in self.tables:
                self.tables.append(p)
                self.table_list.addItem(p)
                new_paths.append(p)
        if new_paths:
            self.log_view.appendPlainText(f"系统：已添加 {len(new_paths)} 个表格。")
            self.log_view.appendPlainText("系统：以下完整路径已同步给AI，您无需在指令中重复填写，供确认参考：")
            for path in new_paths:
                self.log_view.appendPlainText(f"    {path}")
            for path in new_paths:
                self._scan_table_privacy(path)
            self.trigger_intent_recognition()

    def browse_output(self):
        current = self.output_edit.text().strip()
        initial_dir = current
        if current:
            cp = Path(current)
            initial_dir = str(cp.parent if cp.suffix else cp)
        else:
            stored = self.settings.get("last_ai_export_path", "") or ""
            if stored:
                try:
                    initial_dir = str(Path(stored).parent)
                except Exception:
                    initial_dir = str(Path.home())
            else:
                initial_dir = str(Path.home())
        fmt_key = self.export_format_combo.currentData()
        fmt_info = EXPORT_FORMATS.get(fmt_key, EXPORT_FORMATS["excel"])
        path, _ = QFileDialog.getSaveFileName(self, "选择导出文件", initial_dir, fmt_info["dialog_filter"])
        if path:
            p = Path(path)
            if not p.suffix or p.suffix.lower() not in fmt_info["extensions"]:
                p = p.with_suffix(fmt_info["default_extension"])
            self.output_edit.setText(str(p))
            self._sync_format_combo_with_path(str(p))
            self.settings.update(last_ai_export_path=str(p))

    def _get_format_info(self, fmt_key=None):
        key = fmt_key if fmt_key in EXPORT_FORMATS else None
        if key is None:
            current_key = self.export_format_combo.currentData() if hasattr(self, "export_format_combo") else None
            key = current_key if current_key in EXPORT_FORMATS else "excel"
        return EXPORT_FORMATS.get(key, EXPORT_FORMATS["excel"])

    def on_export_format_changed(self, index):
        del index
        fmt_info = self._get_format_info()
        current_path = self.output_edit.text().strip()
        if not current_path:
            return
        p = Path(current_path)
        if p.suffix.lower() not in fmt_info["extensions"]:
            new_path = p.with_suffix(fmt_info["default_extension"])
            self.output_edit.setText(str(new_path))
            self.settings.update(last_ai_export_path=str(new_path))

    def _sync_format_combo_with_path(self, path: str):
        if not hasattr(self, "export_format_combo") or self.export_format_combo is None:
            return
        suffix = Path(path).suffix.lower() if path else ""
        matched_index = None
        for idx in range(self.export_format_combo.count()):
            fmt_key = self.export_format_combo.itemData(idx)
            fmt_info = EXPORT_FORMATS.get(fmt_key, EXPORT_FORMATS["excel"])
            if suffix and suffix in fmt_info["extensions"]:
                matched_index = idx
                break
        self.export_format_combo.blockSignals(True)
        self.export_format_combo.setCurrentIndex(matched_index if matched_index is not None else 0)
        self.export_format_combo.blockSignals(False)

    def show_privacy_statement(self, link=None):
        del link
        dlg = QDialog(self)
        dlg.setWindowTitle("用户隐私声明")
        layout = QVBoxLayout(dlg)
        viewer = QTextBrowser()
        viewer.setMarkdown(PRIVACY_STATEMENT_TEXT)
        viewer.setReadOnly(True)
        viewer.setMinimumSize(560, 360)
        layout.addWidget(viewer)
        btn = QPushButton("关闭")
        btn.clicked.connect(dlg.accept)
        layout.addWidget(btn, alignment=Qt.AlignmentFlag.AlignRight)
        dlg.resize(640, 480)
        dlg.exec()

    def _scan_table_privacy(self, path: str):
        try:
            df = pd.read_excel(path, dtype=str, nrows=200).fillna("")
        except Exception as e:
            self.log_view.appendPlainText(f"系统：无法读取 {path} 进行脱敏检查：{e}")
            return
        _, sensitive_map = desensitize_dataframe(df)
        if sensitive_map:
            cols = ", ".join(sensitive_map.keys())
            self.log_view.appendPlainText(f"系统：检测到敏感字段 {cols}，已自动使用随机样本对AI上下文进行脱敏。")
        else:
            self.log_view.appendPlainText("系统：未检测到明显敏感字段，仍会对上下文样本执行脱敏处理。")

    def append_history(self, role: str, content: str):
        content = (content or "").strip()
        if not content:
            return
        prefix = "👤" if role == "user" else "🤖"
        first_line = content.splitlines()[0]
        if len(first_line) > 60:
            first_line = first_line[:60] + "…"
        item = QListWidgetItem(f"{prefix} {first_line}")
        item.setData(Qt.ItemDataRole.UserRole, content)
        self.history_list.addItem(item)
        self.history_list.scrollToBottom()
        self.conversation_messages.append({"role": role, "content": content})

    def show_history_detail(self, item):
        content = item.data(Qt.ItemDataRole.UserRole) or ""
        dlg = QDialog(self)
        dlg.setWindowTitle("对话详情")
        lay = QVBoxLayout(dlg)
        txt = QPlainTextEdit()
        txt.setReadOnly(True)
        txt.setPlainText(content)
        lay.addWidget(txt)
        btn = QPushButton("关闭")
        btn.clicked.connect(dlg.accept)
        lay.addWidget(btn)
        dlg.resize(520, 320)
        dlg.exec()

    def clear_history(self, checked=False, confirm_dialog=True):
        del checked
        if confirm_dialog and self.conversation_messages:
            confirm = QMessageBox.question(self, "确认", "确定要清空历史对话吗？")
            if confirm != QMessageBox.StandardButton.Yes:
                return
        self.conversation_messages.clear()
        self.history_list.clear()
        self.log_view.appendPlainText("系统：已清空历史对话。")

    def start_new_session(self, checked=False):
        del checked
        if self.worker and self.worker.isRunning():
            confirm = QMessageBox.question(
                self,
                "确认",
                "当前有任务正在执行，是否取消并开启新对话？",
            )
            if confirm != QMessageBox.StandardButton.Yes:
                return
            self.cancel_current()

        self._stop_intent_worker()
        self.log_view.clear()
        self.clear_history(confirm_dialog=False)
        self.tables.clear()
        self.table_list.clear()
        self.message_edit.clear()
        self.code_preview.clear()
        self.table_preview.clear()
        self.table_preview.setRowCount(0)
        self.table_preview.setColumnCount(0)
        self.preview_tabs.setCurrentWidget(self.code_preview)
        self.status_label.setText("等待指令…")
        self._last_status_text = ""
        self.awaiting_execution = False
        self.btn_execute.setEnabled(False)
        self._reset_recommend_placeholder("推荐操作将在此显示")
        self.log_view.appendPlainText("系统：已开启新对话，所有内容已重置。")

    def _stop_intent_worker(self):
        if self.intent_worker and self.intent_worker.isRunning():
            self.intent_worker.requestInterruption()
            self.intent_worker.wait(200)
        self.intent_worker = None

    def _clear_recommend_layout(self):
        while self.recommend_layout.count():
            item = self.recommend_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

    def _reset_recommend_placeholder(self, text="推荐操作将在此显示"):
        self.recommend_buttons.clear()
        self._clear_recommend_layout()
        label = QLabel(text)
        label.setStyleSheet("color: #94a3b8;")
        self.recommend_placeholder = label
        self.recommend_layout.addWidget(label)
        self.recommend_layout.addStretch()

    def trigger_intent_recognition(self):
        if not getattr(self, "recommend_layout", None):
            return
        if not self.tables:
            self._reset_recommend_placeholder("推荐操作将在此显示")
            return

        api_key = (self.settings.get("ai_api_key", "") or "").strip()
        if not api_key:
            self._reset_recommend_placeholder("请在设置中填写API Key以获取推荐。")
            return

        self._reset_recommend_placeholder("正在分析常见操作…")
        self.log_view.appendPlainText("系统：正在分析表格，推荐常见操作…")
        self._stop_intent_worker()
        self.intent_worker = IntentWorker(api_key, self.tables)
        self.intent_worker.results.connect(self.on_intent_results)
        self.intent_worker.error.connect(self.on_intent_error)
        self.intent_worker.finished.connect(self.on_intent_finished)
        self.intent_worker.start()

    def on_intent_results(self, intents: List[str]):
        self.recommend_buttons.clear()
        self._clear_recommend_layout()
        if not intents:
            self._reset_recommend_placeholder("暂未获取推荐操作")
            return

        for text in intents:
            btn = QPushButton(text)
            btn.clicked.connect(lambda _, t=text: self.on_recommendation_clicked(t))
            self.recommend_layout.addWidget(btn)
            self.recommend_buttons.append(btn)

        self.recommend_layout.addStretch()
        self.recommend_placeholder = None
        self.log_view.appendPlainText("系统：推荐操作 → " + "、".join(intents))

    def on_intent_error(self, message: str):
        first_line = (message or "").splitlines()[0] if message else ""
        self._reset_recommend_placeholder("未能获取推荐操作")
        if first_line:
            self.log_view.appendPlainText(f"系统：推荐操作生成失败：{first_line}")

    def on_intent_finished(self):
        self.intent_worker = None

    def on_recommendation_clicked(self, text: str):
        text = text.strip()
        if not text:
            return
        self.message_edit.setPlainText(text)
        self.message_edit.moveCursor(QTextCursor.MoveOperation.End)
        self.send_message()

    def render_table_preview(self, path_str: str):
        self.table_preview.clear()
        self.table_preview.setRowCount(0)
        self.table_preview.setColumnCount(0)
        if not path_str:
            return
        try:
            df = pd.read_excel(path_str)
        except Exception as e:
            self.log_view.appendPlainText(f"系统：无法加载表格预览：{e}")
            return

        preview_df = df.head(200)
        columns = [str(c) for c in preview_df.columns]
        self.table_preview.setColumnCount(len(columns))
        if columns:
            self.table_preview.setHorizontalHeaderLabels(columns)
        self.table_preview.setRowCount(len(preview_df))
        for row_idx, (_, row) in enumerate(preview_df.iterrows()):
            for col_idx, value in enumerate(row):
                display = "" if pd.isna(value) else str(value)
                self.table_preview.setItem(row_idx, col_idx, QTableWidgetItem(display))

        self.table_preview.resizeColumnsToContents()
        self.preview_tabs.setCurrentWidget(self.table_preview)

    def _adopt_result_table(self, path_str: str):
        if not path_str:
            return
        path_str = str(path_str)
        self.tables = [path_str]
        self.table_list.clear()
        self.table_list.addItem(path_str)
        self.log_view.appendPlainText("系统：合并成功，已将生成结果设定为当前处理表格。")
        self.render_table_preview(path_str)
        self.trigger_intent_recognition()

    def send_message(self):
        if self.worker and self.worker.isRunning():
            QMessageBox.warning(self, "提示", "当前有任务正在执行，请稍候。")
            return

        api_key = (self.settings.get("ai_api_key", "") or "").strip()
        if not api_key:
            QMessageBox.warning(self, "提示", "请先在“设置”中填写API Key。")
            return

        if not self.tables:
            QMessageBox.warning(self, "提示", "请至少添加一个表格")
            return

        message = self.message_edit.toPlainText().strip()
        if not message:
            QMessageBox.warning(self, "提示", "请填写对话指令")
            return

        output_path_text = self.output_edit.text().strip()
        if not output_path_text:
            QMessageBox.warning(self, "提示", "请先选择导出结果路径")
            return
        output_path = Path(output_path_text)
        fmt_key = self.export_format_combo.currentData()
        fmt_info = self._get_format_info(fmt_key)
        if not output_path.suffix or output_path.suffix.lower() not in fmt_info["extensions"]:
            output_path = output_path.with_suffix(fmt_info["default_extension"])
            self.output_edit.setText(str(output_path))
            self._sync_format_combo_with_path(str(output_path))
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            QMessageBox.warning(self, "提示", f"无法创建导出目录：{e}")
            return

        self.settings.update(last_ai_export_path=str(output_path))

        temp_map = {
            "代码生成/数学解题": 0.0,
            "数据抽取/分析": 1.0,
            "通用对话": 1.3,
            "翻译": 1.3,
            "创意类写作/诗歌创作": 1.5
        }
        temperature = temp_map.get(self.scenario_combo.currentText(), 0.0)
        language = self.language_combo.currentText()

        history_snapshot = [msg.copy() for msg in self.conversation_messages]
        self.append_history("user", message)
        self.log_view.appendPlainText(f"用户：{message}")
        self.log_view.appendPlainText("")

        self.message_edit.clear()
        self.code_preview.clear()
        self.preview_tabs.setCurrentIndex(0)
        self.status_label.setText("准备中…")
        self._last_status_text = ""
        self.btn_send.setEnabled(False)
        self.btn_execute.setEnabled(False)
        self.awaiting_execution = False

        self.worker = AIWorker(api_key, self.tables, message, temperature, language, str(output_path), fmt_key, history_snapshot)
        self.worker.progress.connect(self.on_worker_progress)
        self.worker.code_stream.connect(self.on_worker_code_stream)
        self.worker.code_ready.connect(self.on_worker_code_ready)
        self.worker.success.connect(self.on_worker_success)
        self.worker.error.connect(self.on_worker_error)
        self.worker.finished.connect(self.on_worker_finished)
        self.worker.start()

    def exec_generated_code(self):
        if not (self.worker and self.worker.isRunning() and self.awaiting_execution):
            return
        self.awaiting_execution = False
        self.btn_execute.setEnabled(False)
        self.status_label.setText("执行中…")
        self.log_view.appendPlainText("系统：开始执行生成的代码。")
        self.worker.approve_execution()

    def cancel_current(self):
        if self.worker and self.worker.isRunning():
            self.worker.requestInterruption()
            self.worker.approve_execution()
            self.status_label.setText("已请求取消…")
            self.log_view.appendPlainText("系统：已请求取消当前任务。")
            self.awaiting_execution = False
            self.btn_execute.setEnabled(False)
        else:
            self.log_view.appendPlainText("系统：当前没有正在执行的任务。")

    def on_worker_progress(self, text: str):
        self.status_label.setText(text)
        if text and text != self._last_status_text:
            self.log_view.appendPlainText(f"系统：{text}")
            self._last_status_text = text

    def on_worker_code_stream(self, text: str):
        self.code_preview.setPlainText(text)
        sb = self.code_preview.verticalScrollBar()
        sb.setValue(sb.maximum())

    def on_worker_code_ready(self, text: str):
        self.on_worker_code_stream(text)
        self.awaiting_execution = True
        self.btn_execute.setEnabled(True)
        self.status_label.setText("代码生成完成，请确认后执行。")
        self.log_view.appendPlainText("系统：模型已生成代码，等待执行。")
        self._last_status_text = "代码生成完成，请确认后执行。"

    def on_worker_success(self, path_str: str):
        msg = f"表格处理已完成，结果保存在：{path_str}。还需要我继续协助处理这个结果吗？"
        self.append_history("assistant", msg)
        self.log_view.appendPlainText(f"成功：{path_str}")
        self.status_label.setText("执行完成")
        self._last_status_text = "执行完成"
        self._adopt_result_table(path_str)
        QMessageBox.information(self, "执行完成", f"已生成文件：\n{path_str}")

    def on_worker_error(self, err: str):
        err = (err or "").strip()
        first_line = err.splitlines()[0] if err else "未知错误"
        self.append_history("assistant", f"执行失败：{first_line}")
        if err:
            self.log_view.appendPlainText("错误：")
            self.log_view.appendPlainText(err)
        self.status_label.setText("执行失败")
        self._last_status_text = "执行失败"
        if first_line != "已取消":
            dlg = QDialog(self)
            dlg.setWindowTitle("错误")
            lay = QVBoxLayout(dlg)
            lay.addWidget(QLabel("执行失败，以下是错误信息："))
            txt = QPlainTextEdit()
            txt.setReadOnly(True)
            txt.setPlainText(err)
            lay.addWidget(txt)
            btn = QPushButton("关闭")
            btn.clicked.connect(dlg.accept)
            lay.addWidget(btn)
            dlg.resize(520, 320)
            dlg.exec()

    def on_worker_finished(self):
        self.worker = None
        self.awaiting_execution = False
        self.btn_execute.setEnabled(False)
        self.btn_send.setEnabled(True)
        self._last_status_text = ""

    def closeEvent(self, event):
        if self.worker and self.worker.isRunning():
            self.cancel_current()
        self._stop_intent_worker()
        super().closeEvent(event)


class SettingsDialog(QDialog):
    def __init__(self, parent, settings: AppSettings):
        super().__init__(parent)
        self.settings = settings
        self.setWindowTitle("设置")
        layout = QVBoxLayout(self)

        layout.addWidget(QLabel("DeepSeek API Key:"))
        self.api_edit = QLineEdit(self.settings.get("ai_api_key", ""))
        self.api_edit.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(self.api_edit)

        layout.addWidget(QLabel("数据报表模块 API Key:"))
        self.report_api_edit = QLineEdit(self.settings.get("report_api_key", ""))
        self.report_api_edit.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(self.report_api_edit)

        layout.addWidget(QLabel("主题模式:"))
        self.theme_combo = QComboBox()
        self.theme_combo.addItems(["深色", "浅色"])
        current_theme = self.settings.get("theme", "dark")
        self.theme_combo.setCurrentIndex(0 if current_theme == "dark" else 1)
        layout.addWidget(self.theme_combo)

        layout.addWidget(QLabel("Excel 写入引擎:"))
        self.engine_combo = QComboBox()
        self.engine_combo.addItems(["自动选择", "仅使用COM", "仅使用openpyxl"])
        engine_mode = self.settings.get("engine_mode", "auto")
        engine_index = {"auto": 0, "com": 1, "openpyxl": 2}.get(engine_mode, 0)
        self.engine_combo.setCurrentIndex(engine_index)
        layout.addWidget(self.engine_combo)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def accept(self):
        theme = "dark" if self.theme_combo.currentIndex() == 0 else "light"
        engine_mode = {0: "auto", 1: "com", 2: "openpyxl"}[self.engine_combo.currentIndex()]
        self.settings.update(
            ai_api_key=self.api_edit.text().strip(),
            report_api_key=self.report_api_edit.text().strip(),
            theme=theme,
            engine_mode=engine_mode
        )
        super().accept()


def openpyxl_write_and_save_optimized(tgt_path: Path, tgt_sheet: str, out_path: Path,
                                    df_src: pd.DataFrame, df_tgt: pd.DataFrame, src_map: pd.DataFrame,
                                    mapping: list, tgt_field_to_col: dict, tgt_data_start_row: int,
                                    overwrite_all: bool):
    """【兼容核心】openpyxl 版本的写入，不依赖Windows Office"""
    wb = None
    total_found, total_write = 0, 0
    
    try:
        wb = load_workbook(tgt_path, data_only=False, read_only=False, keep_vba=True)
        ws = wb[tgt_sheet]
        
        updates_by_col = {}
        for row_offset, (_, row) in enumerate(df_tgt.iterrows()):
            idx_val = row.get("_IDX_", "")
            if not idx_val or idx_val not in src_map.index: continue
            
            src_row = src_map.loc[idx_val]
            total_found += 1
            excel_row = tgt_data_start_row + row_offset
            
            for tgt_field, src_field in mapping:
                if not src_field or src_field not in df_src.columns: continue
                tgt_col = tgt_field_to_col.get(tgt_field)
                if not tgt_col: continue
                
                val = src_row.get(src_field)
                if val == '' or pd.isna(val): continue
                
                cell = ws.cell(row=excel_row, column=tgt_col)
                if not overwrite_all and cell.value is not None and str(cell.value).strip() != "": continue
                
                if tgt_col not in updates_by_col: updates_by_col[tgt_col] = []
                updates_by_col[tgt_col].append((excel_row, val))
        
        for tgt_col, updates in updates_by_col.items():
            for excel_row, val in updates:
                cell = ws.cell(row=excel_row, column=tgt_col)
                fmt = {
                    'font': copy(cell.font),
                    'fill': copy(cell.fill),
                    'border': copy(cell.border),
                    'alignment': copy(cell.alignment),
                    'number_format': cell.number_format,
                    'protection': copy(cell.protection)
                }
                cell.value = val
                cell.font = fmt['font']
                cell.fill = fmt['fill']
                cell.border = fmt['border']
                cell.alignment = fmt['alignment']
                cell.number_format = fmt['number_format']
                cell.protection = fmt['protection']
                total_write += 1
        
        wb.save(out_path)
        
    finally:
        if wb:
            try: wb.close()
            except: pass
        gc.collect()
    
    return total_found, total_write

# ===================== 数据匹配模块 =====================

class DataMatchingWidget(QWidget):
    def __init__(self, settings: AppSettings, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.settings = settings
        # 多表模式下，源表和目标表都可能有多个，通过列表维护
        self.src_groups, self.tgt_groups = [], []
        self.src_headers, self.tgt_headers = [], []
        self.mode = "one2one"  # 默认一对一
        self._init_ui()

    def _init_ui(self):
        layout = QGridLayout(self)
        layout.setContentsMargins(14, 12, 14, 12); layout.setSpacing(12)

        # ------ 模式选择 ------
        mode_box = QGroupBox("匹配模式")
        mode_layout = QHBoxLayout(mode_box)
        self.rb_one2one = QRadioButton("一对一"); self.rb_one2one.setChecked(True)
        self.rb_one2many = QRadioButton("一对多")
        self.rb_many2one = QRadioButton("多对一")
        mode_layout.addWidget(self.rb_one2one); mode_layout.addWidget(self.rb_one2many); mode_layout.addWidget(self.rb_many2one)
        self.rb_one2one.toggled.connect(self.on_mode_change)
        self.rb_one2many.toggled.connect(self.on_mode_change)
        self.rb_many2one.toggled.connect(self.on_mode_change)
        layout.addWidget(mode_box, 0, 0, 1, 2)

        # ------ 动态源/目标面板 ------
        self.src_container = QWidget(); self.src_layout = QVBoxLayout(self.src_container)
        self.src_layout.setContentsMargins(0,0,0,0); self.src_layout.setSpacing(6)
        self.tgt_container = QWidget(); self.tgt_layout = QVBoxLayout(self.tgt_container)
        self.tgt_layout.setContentsMargins(0,0,0,0); self.tgt_layout.setSpacing(6)

        self.btn_add_src = QPushButton("添加信息源"); self.btn_add_src.clicked.connect(self.add_source_group)
        self.btn_add_tgt = QPushButton("添加目标表"); self.btn_add_tgt.clicked.connect(self.add_target_group)
        self.src_tabs = QTabWidget(); self.tgt_tabs = QTabWidget()
        self.src_layout.addWidget(self.btn_add_src); self.src_layout.addWidget(self.src_tabs)
        self.tgt_layout.addWidget(self.btn_add_tgt); self.tgt_layout.addWidget(self.tgt_tabs)

        layout.addWidget(self.src_container, 1, 0)
        layout.addWidget(self.tgt_container, 1, 1)

        # 初始各添加一个组
        self.add_source_group()
        self.add_target_group()

        map_grp = QGroupBox("字段映射与执行")
        map_layout = QVBoxLayout(map_grp)
        self.map_table = QTableWidget(0, 2)
        self.map_table.setHorizontalHeaderLabels(["目标字段（写入此列）", "来自信息源（下拉选择/跳过）"])
        self.map_table.horizontalHeader().setStretchLastSection(True)
        self.map_table.verticalHeader().setVisible(False)
        self.map_table.setEditTriggers(QAbstractItemView.EditTrigger.AllEditTriggers)
        
        opts_layout = QHBoxLayout()
        opts_layout.addWidget(QLabel("源索引:")); self.cmb_src_index = QComboBox(); opts_layout.addWidget(self.cmb_src_index, 1)
        opts_layout.addWidget(QLabel("目标索引:")); self.cmb_tgt_index = QComboBox(); opts_layout.addWidget(self.cmb_tgt_index, 1)

        write_mode_box = QGroupBox("写入模式"); write_mode_layout = QHBoxLayout(write_mode_box)
        self.rb_fill_empty = QRadioButton("仅填充空值"); self.rb_fill_empty.setChecked(True)
        self.rb_overwrite = QRadioButton("覆盖所有值")
        write_mode_layout.addWidget(self.rb_fill_empty); write_mode_layout.addWidget(self.rb_overwrite)
        opts_layout.addWidget(write_mode_box, 2)

        btns = QHBoxLayout()
        self.btn_load_config = QPushButton("加载方案"); self.btn_save_config = QPushButton("保存方案")
        self.btn_auto = QPushButton("自动预填"); self.btn_run = QPushButton("执行匹配并导出")
        btns.addWidget(self.btn_load_config); btns.addWidget(self.btn_save_config)
        btns.addStretch(); btns.addWidget(self.btn_auto); btns.addWidget(self.btn_run)
        
        map_layout.addWidget(self.map_table); map_layout.addLayout(opts_layout); map_layout.addLayout(btns)
        layout.addWidget(map_grp, 2, 0, 1, 2)

        self.btn_load_config.clicked.connect(self.load_mapping_config)
        self.btn_save_config.clicked.connect(self.save_mapping_config)
        self.btn_auto.clicked.connect(self.auto_fill_mapping)
        self.btn_run.clicked.connect(self.run_and_export)
        self.on_mode_change()

    # ---- 动态增减源/目标组 ----
    def add_source_group(self):
        idx = len(self.src_groups) + 1
        g = self._build_config_group(f"信息源{idx}", is_source=True)
        self.src_groups.append(g)
        self.src_tabs.addTab(g, f"信息源{idx}")
        self.src_tabs.setCurrentWidget(g)

    def add_target_group(self):
        idx = len(self.tgt_groups) + 1
        g = self._build_config_group(f"目标表{idx}", is_source=False)
        self.tgt_groups.append(g)
        self.tgt_tabs.addTab(g, f"目标表{idx}")
        self.tgt_tabs.setCurrentWidget(g)

    def on_mode_change(self):
        if self.rb_one2one.isChecked():
            self.mode = "one2one"
        elif self.rb_one2many.isChecked():
            self.mode = "one2many"
        else:
            self.mode = "many2one"

        self.btn_add_src.setEnabled(self.mode == "many2one")
        self.btn_add_tgt.setEnabled(self.mode == "one2many")

        if self.mode != "many2one":
            while len(self.src_groups) > 1:
                g = self.src_groups.pop()
                idx = self.src_tabs.count() - 1
                self.src_tabs.removeTab(idx)
                g.deleteLater()
        if self.mode != "one2many":
            while len(self.tgt_groups) > 1:
                g = self.tgt_groups.pop()
                idx = self.tgt_tabs.count() - 1
                self.tgt_tabs.removeTab(idx)
                g.deleteLater()

        self._recalc_headers()

    def _recalc_headers(self):
        self.src_headers = []
        for g in self.src_groups:
            if hasattr(g, "_headers"):
                for h in g._headers:
                    if h not in self.src_headers:
                        self.src_headers.append(h)
        self.cmb_src_index.clear()
        if self.src_headers:
            self.cmb_src_index.addItems(self.src_headers)

        self.tgt_headers = []
        if self.tgt_groups and hasattr(self.tgt_groups[0], "_headers"):
            self.tgt_headers = self.tgt_groups[0]._headers
            self.cmb_tgt_index.clear(); self.cmb_tgt_index.addItems(self.tgt_headers)

        self.rebuild_mapping_table()

    def _build_config_group(self, title, is_source: bool):
        g = QWidget(); grid = QGridLayout(g)
        grid.setContentsMargins(6,6,6,6)
        grid.setHorizontalSpacing(6)
        grid.setVerticalSpacing(4)
        le_path = QLineEdit(); le_path.setReadOnly(True)
        btn_browse = QPushButton("浏览…")
        cmb_sheet = QComboBox()
        sp_header = QSpinBox(); sp_header.setRange(1, 100000); sp_header.setValue(1)
        sp_startcol = QSpinBox(); sp_startcol.setRange(1, 10000); sp_startcol.setValue(1)
        btn_extract = QPushButton("提取字段并预览")
        preview_table = QTableWidget(); preview_table.setRowCount(5); preview_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        
        grid.addWidget(QLabel("Excel 文件："), 0, 0); grid.addWidget(le_path, 0, 1); grid.addWidget(btn_browse, 0, 2)
        grid.addWidget(QLabel("工作表："), 1, 0); grid.addWidget(cmb_sheet, 1, 1, 1, 2)
        grid.addWidget(QLabel("表头行："), 2, 0); grid.addWidget(sp_header, 2, 1)
        grid.addWidget(QLabel("数据起始列："), 2, 2); grid.addWidget(sp_startcol, 2, 3)
        grid.addWidget(btn_extract, 3, 0, 1, 4); grid.addWidget(preview_table, 4, 0, 1, 4)

        def auto_detect(_=None):
            p = le_path.text().strip()
            sheet = cmb_sheet.currentText().strip()
            if not p or not sheet:
                return
            try:
                h, c = auto_detect_header_start(Path(p), sheet)
                sp_header.setValue(h)
                sp_startcol.setValue(c)
                self._update_fields_and_preview(g, is_source)
            except Exception:
                pass

        def on_browse():
            path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel (*.xlsx *.xlsm)")
            if not path: return
            le_path.setText(path)
            try:
                sheet_names = get_sheet_names_safely(Path(path))
                cmb_sheet.clear(); cmb_sheet.addItems(sheet_names)
                auto_detect()
            except Exception as e: QMessageBox.critical(self, "错误", f"无法读取工作表：\n{e}")
        btn_browse.clicked.connect(on_browse)

        cmb_sheet.currentTextChanged.connect(auto_detect)

        def on_extract(): self._update_fields_and_preview(g, is_source)
        btn_extract.clicked.connect(on_extract)
        
        g._le_path, g._cmb_sheet, g._sp_header, g._sp_startcol, g._preview_table = le_path, cmb_sheet, sp_header, sp_startcol, preview_table
        return g

    def _update_fields_and_preview(self, group_box, is_source):
        p_str = group_box._le_path.text().strip()
        if not p_str: QMessageBox.warning(self, "提示", "请先选择Excel文件。"); return
        path, sheet = Path(p_str), group_box._cmb_sheet.currentText().strip()
        if not sheet: QMessageBox.warning(self, "提示", "请先选择工作表。"); return
        header_row, start_col = group_box._sp_header.value(), group_box._sp_startcol.value()

        try:
            df = read_excel_dataframe(path, sheet, header_row, start_col)
            headers = list(df.columns)
            self._update_preview_table(group_box._preview_table, df)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"提取字段失败：\n{e}"); return

        group_box._path, group_box._sheet = path, sheet
        group_box._header_row, group_box._start_col = header_row, start_col
        group_box._headers = headers

        if (not is_source) and self.tgt_groups and group_box != self.tgt_groups[0]:
            if hasattr(self.tgt_groups[0], "_headers") and headers != self.tgt_groups[0]._headers:
                QMessageBox.warning(self, "警告", "目标表字段结构不一致，可能导致匹配问题。")

        self._recalc_headers()

        if is_source and self.src_headers:
            if (guess := suggest_index_choice(self.src_headers)):
                self.cmb_src_index.setCurrentText(guess)
        if (not is_source) and self.tgt_headers:
            if (guess := suggest_index_choice(self.tgt_headers)):
                self.cmb_tgt_index.setCurrentText(guess)

    def _update_preview_table(self, table: QTableWidget, df: pd.DataFrame):
        preview_df = df.head(5)
        table.clear()
        table.setColumnCount(len(preview_df.columns))
        table.setHorizontalHeaderLabels(preview_df.columns)
        table.setRowCount(len(preview_df))
        for r_idx, row in enumerate(preview_df.itertuples(index=False)):
            for c_idx, val in enumerate(row):
                table.setItem(r_idx, c_idx, QTableWidgetItem(str(val)))
        table.resizeColumnsToContents()

    def rebuild_mapping_table(self):
        self.map_table.clearContents()
        self.map_table.setRowCount(len(self.tgt_headers))
        options = ["<跳过>"] + self.src_headers
        delegate = ComboDelegate(self.map_table, options)
        self.map_table.setItemDelegateForColumn(1, delegate)
        
        for r, tgt_name in enumerate(self.tgt_headers):
            item0 = QTableWidgetItem(tgt_name)
            item0.setFlags(item0.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.map_table.setItem(r, 0, item0)
            self.map_table.setItem(r, 1, QTableWidgetItem("<跳过>"))
        self.map_table.resizeColumnsToContents()

    def auto_fill_mapping(self):
        """【核心改进】使用模糊匹配自动填充，并提供视觉反馈"""
        if not self.src_headers: QMessageBox.warning(self, "提示", "请先提取信息源的字段。"); return

        unmatched_color = QBrush(QColor("#111827")) # 默认背景色
        matched_color = QBrush(QColor("#14532d"))   # 匹配成功的背景色

        for r in range(self.map_table.rowCount()):
            tgt_name = self.map_table.item(r, 0).text()
            best_match = find_best_match(tgt_name, self.src_headers)
            
            item_tgt = self.map_table.item(r, 0)
            item_src = self.map_table.item(r, 1)

            if best_match:
                item_src.setText(best_match)
                item_tgt.setBackground(matched_color)
                item_src.setBackground(matched_color)
            else:
                item_src.setText("<跳过>")
                item_tgt.setBackground(unmatched_color)
                item_src.setBackground(unmatched_color)

    def save_mapping_config(self):
        if self.mode != "one2one":
            QMessageBox.information(self, "提示", "仅一对一模式支持保存方案。")
            return

        g_src, g_tgt = self.src_groups[0], self.tgt_groups[0]
        if not hasattr(g_src, "_path") or not hasattr(g_tgt, "_path"):
            QMessageBox.warning(self, "提示", "请先配置并提取信息源和目标表的字段。"); return

        mapping = {self.map_table.item(r, 0).text(): self.map_table.item(r, 1).text() for r in range(self.map_table.rowCount())}
        config = {
            "source": {"path": str(g_src._path), "sheet": g_src._sheet, "header_row": g_src._header_row, "start_col": g_src._start_col},
            "target": {"path": str(g_tgt._path), "sheet": g_tgt._sheet, "header_row": g_tgt._header_row, "start_col": g_tgt._start_col},
            "indices": {"source": self.cmb_src_index.currentText(), "target": self.cmb_tgt_index.currentText()},
            "write_mode": "overwrite" if self.rb_overwrite.isChecked() else "fill_empty", "mapping": mapping
        }
        path, _ = QFileDialog.getSaveFileName(self, "保存映射方案", "", "JSON Files (*.json)")
        if path:
            try:
                with open(path, 'w', encoding='utf-8') as f: json.dump(config, f, indent=4, ensure_ascii=False)
                QMessageBox.information(self, "成功", f"映射方案已保存至:\n{path}")
            except Exception as e: QMessageBox.critical(self, "错误", f"保存失败：\n{e}")

    def load_mapping_config(self):
        if self.mode != "one2one":
            QMessageBox.information(self, "提示", "仅在一对一模式下可加载方案。")
            return

        path, _ = QFileDialog.getOpenFileName(self, "加载映射方案", "", "JSON Files (*.json)")
        if not path: return
        try:
            with open(path, 'r', encoding='utf-8') as f: config = json.load(f)

            for g, cfg in [(self.src_groups[0], config["source"]), (self.tgt_groups[0], config["target"])]:
                g._le_path.setText(cfg["path"])
                g._sp_header.setValue(cfg["header_row"])
                g._sp_startcol.setValue(cfg["start_col"])
                try:
                    sheet_names = get_sheet_names_safely(Path(cfg["path"]))
                    g._cmb_sheet.clear(); g._cmb_sheet.addItems(sheet_names)
                    g._cmb_sheet.setCurrentText(cfg["sheet"])
                except Exception as e: QMessageBox.warning(self, "警告", f"无法加载工作表 {cfg['path']}: {e}")

            self._update_fields_and_preview(self.src_groups[0], True)
            self._update_fields_and_preview(self.tgt_groups[0], False)

            self.cmb_src_index.setCurrentText(config["indices"]["source"])
            self.cmb_tgt_index.setCurrentText(config["indices"]["target"])
            self.rb_overwrite.setChecked(config["write_mode"] == "overwrite")

            for r in range(self.map_table.rowCount()):
                tgt_field = self.map_table.item(r, 0).text()
                src_field = config["mapping"].get(tgt_field, "<跳过>")
                self.map_table.item(r, 1).setText(src_field)

        except Exception as e: QMessageBox.critical(self, "错误", f"加载方案失败：\n{e}")

    def run_and_export(self):
        self.setEnabled(False)
        self.btn_run.setText("正在匹配...")
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        QApplication.processEvents()
        try: 
            self._execute_matching_logic()
        finally:
            QApplication.restoreOverrideCursor()
            self.btn_run.setText("执行匹配并导出")
            self.setEnabled(True)

    def _execute_matching_logic(self):
        if not (self.src_groups and self.tgt_groups and self.src_headers and self.tgt_headers):
            QMessageBox.warning(self, "提示", "请先提取信息源和目标表的字段。"); return
        src_idx, tgt_idx = self.cmb_src_index.currentText(), self.cmb_tgt_index.currentText()
        if not src_idx or not tgt_idx:
            QMessageBox.warning(self, "提示", "请选择索引字段。"); return

        mapping = [(self.map_table.item(r,0).text(), self.map_table.item(r,1).text()) for r in range(self.map_table.rowCount())]
        mapping = [(t, s) for t, s in mapping if s != "<跳过>"]

        # ---- 构建源数据 ----
        try:
            if self.mode == "many2one":
                dfs = [read_excel_dataframe(g._path, g._sheet, g._header_row, g._start_col, True) for g in self.src_groups]
                df_src = pd.concat(dfs, ignore_index=True)
            else:
                g = self.src_groups[0]
                df_src = read_excel_dataframe(g._path, g._sheet, g._header_row, g._start_col, True)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取信息源失败：\n{e}"); return

        if src_idx not in df_src.columns:
            QMessageBox.critical(self, "错误", f"源表无索引：{src_idx}"); return
        df_src["_IDX_"] = df_src[src_idx].apply(norm_str)
        src_map = df_src.drop_duplicates(subset=["_IDX_"], keep='last').set_index("_IDX_")

        engine_mode = self.settings.get("engine_mode", "auto")
        results = []
        for tgt_g in self.tgt_groups:
            try:
                df_tgt = read_excel_dataframe(tgt_g._path, tgt_g._sheet, tgt_g._header_row, tgt_g._start_col, False)
            except Exception as e:
                QMessageBox.critical(self, "错误", f"读取目标表失败：\n{e}"); return
            if tgt_idx not in df_tgt.columns:
                QMessageBox.critical(self, "错误", f"目标表无索引：{tgt_idx}"); return
            df_tgt["_IDX_"] = df_tgt[tgt_idx].apply(norm_str)
            tgt_field_to_col = {name: i + tgt_g._start_col for i, name in enumerate(self.tgt_headers)}
            overwrite_all = self.rb_overwrite.isChecked()
            out_path = Path(tgt_g._path).with_name(f"{Path(tgt_g._path).stem}_匹配输出{Path(tgt_g._path).suffix}")

            engine = ""
            errors = {}
            if engine_mode in ("auto", "com"):
                try:
                    total_found, total_write = excel_com_write_and_save_optimized(
                        tgt_g._path, tgt_g._sheet, out_path, df_src, df_tgt, src_map,
                        mapping, tgt_field_to_col, tgt_g._header_row + 1, overwrite_all)
                    engine = "Excel COM（批量优化）"
                except Exception as e1:
                    errors["com"] = e1
                    if engine_mode == "com":
                        QMessageBox.critical(self, "错误", f"COM 保存失败：\n{e1}")
                        return
            if not engine and engine_mode in ("auto", "openpyxl"):
                try:
                    total_found, total_write = openpyxl_write_and_save_optimized(
                        tgt_g._path, tgt_g._sheet, out_path, df_src, df_tgt, src_map,
                        mapping, tgt_field_to_col, tgt_g._header_row + 1, overwrite_all)
                    engine = "openpyxl（Office365兼容）"

                    # 【关键修复】使用'replace'模式安全地进行Pandas验证回写
                    try:
                        df_verify = pd.read_excel(out_path, sheet_name=tgt_g._sheet, dtype=str).fillna('')
                        with pd.ExcelWriter(
                            out_path,
                            engine='openpyxl',
                            mode='a',
                            if_sheet_exists='replace'
                        ) as writer:
                            df_verify.to_excel(writer, sheet_name=tgt_g._sheet, index=False)
                        engine = "openpyxl（Pandas兼容性优化）"
                    except Exception:
                        # 验证失败，但原文件可能仍然可用
                        pass
                except Exception as e2:
                    errors["openpyxl"] = e2
                    if engine_mode == "openpyxl":
                        QMessageBox.critical(self, "错误", f"openpyxl 保存失败：\n{e2}")
                        return
            if not engine:
                def _short(err, default):
                    if err is None:
                        return default
                    msg = str(err)
                    return msg if len(msg) <= 200 else msg[:200] + "..."

                com_default = "未尝试（根据设置跳过）" if engine_mode == "openpyxl" else "无可用错误信息"
                op_default = "未尝试（根据设置跳过）" if engine_mode == "com" else "无可用错误信息"
                com_msg = _short(errors.get("com"), com_default)
                op_msg = _short(errors.get("openpyxl"), op_default)
                QMessageBox.critical(
                    self,
                    "错误",
                    f"所有保存方式均失败：\n\nCOM错误：{com_msg}\n\nopenpyxl错误：{op_msg}\n\n建议：\n1. 确保目标Excel文件未被其他程序占用\n2. 检查文件权限\n3. 尝试关闭Excel程序后重试"
                )
                return

            results.append((out_path, engine, total_found, total_write))

        if not results:
            return
        if len(results) == 1:
            out_path, engine, total_found, total_write = results[0]
            QMessageBox.information(self, "完成",
                f"匹配完成（引擎：{engine}）：\n\n"
                f"命中索引记录： {total_found}\n"
                f"共写入单元格： {total_write}\n\n"
                f"结果已导出至：\n{out_path}")
        else:
            msg = "\n\n".join([f"{p}: 命中{f} 写入{w}" for p, _, f, w in results])
            QMessageBox.information(self, "完成", f"已处理{len(results)}个目标表：\n\n{msg}")


# ===================== 数据报表与清洗模块 =====================


class ReportGenerationWorker(QThread):
    progress = pyqtSignal(str)
    error = pyqtSignal(str)
    success = pyqtSignal(str, str)
    code_ready = pyqtSignal(str)

    def __init__(self, api_key: str, sample_csv: str, columns: List[str], instruction: str, dataframe: pd.DataFrame):
        super().__init__()
        self.api_key = api_key.strip()
        self.sample_csv = sample_csv
        self.columns = columns
        self.instruction = instruction.strip()
        self.dataframe = dataframe

    def _prepare_prompt(self) -> tuple[str, str]:
        column_text = ", ".join(map(str, self.columns)) or "(无列信息)"
        preview_text = self.sample_csv.strip() or "(空表)"
        rows, cols = self.dataframe.shape
        fingerprint_text = generate_dataframe_fingerprint(self.dataframe)
        system_prompt = textwrap.dedent(
            """
            你是一位资深的数据分析师与前端可视化工程师。你的任务是读取结构化数据后，生成信息丰富且具有响应式布局的 HTML 数据分析报告。
            编写 Python 脚本时请严格遵循以下约束：
            1. 仅允许导入 pandas、numpy（如需）、json、math、statistics、datetime、textwrap 及其他 Python 标准库；严禁使用 pyecharts、matplotlib 或任何额外的第三方可视化库。
            2. 必须从环境变量 REPORT_CSV_PATH 指向的 UTF-8 CSV 文件加载完整数据集，并在任何计算前通过 pd.to_numeric(..., errors="coerce") 或 pd.to_datetime(..., errors="coerce") 做好类型清洗，同时妥善处理缺失值。
            3. 生成的 HTML 需包含 <html>、<head>、<body> 等完整结构，在 <head> 中编写内联 CSS，确保布局在桌面端与移动端均表现良好（例如限制内容最大宽度、使用 flex 栅格、响应式卡片等）。
            4. 所有图表必须使用原生 ECharts JavaScript 库渲染，在 HTML 中引入 <script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>，并由 Python 将数据转换为 JSON 串嵌入脚本，使用 echarts.init(...) 和 setOption(...) 完成渲染。
            5. 每个图表初始化后需注册 window.addEventListener('resize', chart.resize) 保持响应式，图表容器 div 需设置明确高度并自动适应父容器宽度。
            6. 除图表外，请提供文字洞察、关键指标卡片或汇总表格等辅助信息，帮助读者快速理解数据结论。
            7. 若某个洞察所需字段在清洗后为空，可构造少量示例数据用于演示，并在标题或描述中明确标注“示例数据”；真实数据可用时必须优先展示真实结果。
            8. 最终脚本必须且只能 print(完整 HTML 字符串)，禁止写入文件或输出其他调试信息。
            9. 返回的内容必须是纯 Python 源码，严禁包含 Markdown 代码块标记（例如 ```python）或额外说明文字。
            """
        ).strip()
        user_prompt = textwrap.dedent(
            f"""
            [数据概览]
            - 列名: {column_text}
            - 总行数: {rows}
            - 总列数: {cols}

            [脱敏样本 CSV]
            {preview_text}

            [数据画像]
            {fingerprint_text}

            [分析需求]
            {self.instruction or '请基于该数据集完成一次全面的探索性分析，找出主要趋势、异常与改进建议。'}

            交付要求：
            - 至少构建三个核心洞察点，并为每个洞察提供相应的文字说明。
            - 生成不少于两个的 ECharts 可视化图表（如趋势、对比或占比），保证容器在不同屏幕尺寸下的可读性。
            - HTML 正文需包含主标题、概述段落、要点列表以及总结部分。

            请在遵守系统提示的全部约束下生成 Python 脚本，并仅打印最终 HTML 字符串。
            """
        ).strip()
        return system_prompt, user_prompt

    def run(self):
        if not self.api_key:
            self.error.emit("未配置数据报表模块 API Key")
            return

        try:
            from openai import OpenAI
        except Exception as e:
            self.error.emit(f"未安装openai库: {e}")
            return

        system_prompt, user_prompt = self._prepare_prompt()
        client = OpenAI(api_key=self.api_key, base_url="https://api.deepseek.com/v1")

        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ]

        last_error = ""
        max_attempts = 3

        with tempfile.TemporaryDirectory() as td:
            data_path = Path(td) / "dataset.csv"
            try:
                self.dataframe.to_csv(data_path, index=False, encoding="utf-8")
            except Exception as e:
                self.error.emit(f"写入临时数据失败：{e}")
                return

            env = os.environ.copy()
            env.setdefault("PYTHONPATH", env.get("PYTHONPATH", ""))
            env["REPORT_CSV_PATH"] = str(data_path)
            env["PYTHONIOENCODING"] = "utf-8"

            script_path = Path(td) / "report.py"
            latest_code = ""

            for attempt in range(max_attempts):
                if attempt == 0:
                    self.progress.emit("正在调用模型生成报告代码…")
                else:
                    self.progress.emit(f"正在请求模型修复脚本（第{attempt + 1}次尝试）…")

                try:
                    response = client.chat.completions.create(
                        model="deepseek-chat",
                        temperature=0.2,
                        messages=messages,
                    )
                except Exception as e:
                    self.error.emit(str(e))
                    return

                content = ""
                try:
                    content = (response.choices[0].message.content or "").strip()
                except Exception:
                    self.error.emit("模型返回内容为空")
                    return

                if content.startswith("```"):
                    parts = content.splitlines()
                    content = "\n".join(parts[1:-1]) if len(parts) >= 2 else content

                code = normalize_generated_code(content.strip())
                if not code:
                    self.error.emit("未获取到有效的代码内容")
                    return

                latest_code = code
                self.code_ready.emit(code)

                try:
                    script_path.write_text(code, encoding="utf-8")
                except Exception as e:
                    self.error.emit(f"写入生成脚本失败：{e}")
                    return

                self.progress.emit("正在执行生成的分析脚本…")
                try:
                    proc = subprocess.run(
                        [sys.executable, str(script_path)],
                        capture_output=True,
                        text=True,
                        encoding="utf-8",
                        errors="replace",
                        cwd=td,
                        env=env,
                        timeout=180,
                    )
                except Exception as e:
                    last_error = f"执行脚本失败：{e}"
                    error_for_prompt = last_error
                else:
                    stdout = (proc.stdout or "").strip()
                    stderr = (proc.stderr or "").strip()
                    if proc.returncode == 0 and stdout:
                        self.success.emit(stdout, latest_code)
                        return
                    if proc.returncode == 0:
                        last_error = "脚本执行成功但未返回HTML内容"
                    else:
                        last_error = stderr or stdout or "未知错误"
                    error_for_prompt = stderr or stdout or last_error

                if attempt == max_attempts - 1:
                    break

                condensed_error = summarize_error(error_for_prompt, self.columns)
                prompt_error = error_for_prompt if len(error_for_prompt) <= 1500 else error_for_prompt[:1500] + "…"
                messages.append({"role": "assistant", "content": f"```python\n{code}\n```"})
                messages.append({
                    "role": "user",
                    "content": (
                        "上述代码执行失败。错误摘要："
                        f"{condensed_error}\n详细日志：{prompt_error}\n"
                        "请在保留既有需求的基础上修复代码，并重新给出完整的可运行Python脚本。"
                    ),
                })

        self.error.emit(f"报告生成失败：{last_error or '未知错误'}")


class ScraperWorker(QThread):
    progress = pyqtSignal(str)
    error = pyqtSignal(str)
    data_ready = pyqtSignal(object, str)

    def __init__(self, api_key: str, url: str, instruction: str):
        super().__init__()
        self.api_key = api_key.strip()
        self.url = url.strip()
        self.instruction = instruction.strip()

    def run(self):
        if not self.api_key:
            self.error.emit("未配置数据报表模块 API Key")
            return
        if not self.url:
            self.error.emit("请输入有效的网址")
            return
        if not self.instruction:
            self.error.emit("请输入爬取需求")
            return

        try:
            from openai import OpenAI
        except Exception as e:
            self.error.emit(f"未安装openai库: {e}")
            return

        try:
            import requests
        except Exception as e:
            self.error.emit(f"未安装requests库: {e}")
            return

        self.progress.emit("正在抓取网页结构信息…")
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
                "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            }
            response = requests.get(self.url, headers=headers, timeout=20)
            response.raise_for_status()
            html_text = response.text
        except Exception as e:
            self.error.emit(f"获取网页HTML失败：{e}")
            return

        max_html_chars = 40000
        html_excerpt = html_text[:max_html_chars]
        if len(html_text) > max_html_chars:
            html_excerpt += "\n<!-- HTML内容已截断，仅展示前40000字符 -->"

        system_prompt = textwrap.dedent(
            """
            你是一名资深的网络爬虫工程师，需要根据页面结构编写可直接运行的 Python 脚本来提取结构化数据。
            编写脚本时请严格遵循以下规范：
            1. 仅可使用 requests、os、sys、time、re、json、pandas 以及 bs4 中的 BeautifulSoup（含 Python 标准库）；禁止安装或导入其他第三方库。
            2. 必须从环境变量 SCRAPER_TARGET_URL 读取目标地址，使用 requests.get 发起请求，设置常见浏览器 User-Agent、合理的超时与异常处理。必要时在失败后打印错误并安全退出。
            3. 使用 BeautifulSoup('html.parser' 或 'lxml') 对响应进行解析，准确定位用户需求涉及的 DOM 节点，并在提取每个字段前判断元素是否存在或使用 try/except 捕获异常。
            4. 将提取的记录存入列表后构建 pandas.DataFrame，并通过 df.to_csv(os.environ['SCRAPER_OUTPUT_PATH'], index=False, encoding='utf-8-sig') 写入结果。
            5. 可以使用适量的 print 输出执行进度或结果统计，但禁止输出 Markdown、堆栈跟踪或与任务无关的信息。
            6. 返回的内容必须是纯 Python 源码，绝不允许出现 Markdown 代码块标记（例如 ```python）或说明性段落；建议封装 main() 函数并以 if __name__ == "__main__": main() 作为入口。
            """
        ).strip()

        user_prompt = textwrap.dedent(
            f"""
            [目标网址]
            {self.url}

            [用户任务]
            {self.instruction}

            [页面 HTML 快照（截取）]
            {html_excerpt}

            请仔细阅读上述 DOM 结构，明确数据所在标签和类名后再设计解析逻辑。
            交付的脚本必须：
            - 调用 requests.get 采集 SCRAPER_TARGET_URL，必要时在失败时退出并输出友好的错误信息；
            - 使用 BeautifulSoup 遍历所有匹配条目，逐字段提取文本并在缺失时填充为空字符串；
            - 汇总为 pandas.DataFrame 并写入 CSV 文件（路径由 SCRAPER_OUTPUT_PATH 指定），最后打印成功行数或提示语。

            请输出满足以上要求的完整 Python 代码，且务必不要包含 Markdown 标记。
            """
        ).strip()

        client = OpenAI(api_key=self.api_key, base_url="https://api.deepseek.com/v1")

        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ]

        last_error = ""
        max_attempts = 3

        with tempfile.TemporaryDirectory() as td:
            output_path = Path(td) / "scraped.csv"
            script_path = Path(td) / "scraper.py"

            env = os.environ.copy()
            env.setdefault("PYTHONPATH", env.get("PYTHONPATH", ""))
            env["SCRAPER_TARGET_URL"] = self.url
            env["SCRAPER_OUTPUT_PATH"] = str(output_path)
            env["PYTHONIOENCODING"] = "utf-8"

            for attempt in range(max_attempts):
                if attempt == 0:
                    self.progress.emit("正在生成爬虫脚本…")
                else:
                    self.progress.emit(f"正在请求模型修复脚本（第{attempt + 1}次尝试）…")

                try:
                    response = client.chat.completions.create(
                        model="deepseek-chat",
                        temperature=0.2,
                        messages=messages,
                    )
                except Exception as e:
                    self.error.emit(str(e))
                    return

                content = ""
                try:
                    content = (response.choices[0].message.content or "").strip()
                except Exception:
                    self.error.emit("模型返回内容为空")
                    return

                if content.startswith("```"):
                    parts = content.splitlines()
                    content = "\n".join(parts[1:-1]) if len(parts) >= 2 else content

                code = normalize_generated_code(content.strip())
                if not code:
                    self.error.emit("未获取到有效的爬虫代码")
                    return

                try:
                    script_path.write_text(code, encoding="utf-8")
                except Exception as e:
                    self.error.emit(f"写入脚本失败：{e}")
                    return

                self.progress.emit("正在执行爬虫脚本…")
                try:
                    proc = subprocess.run(
                        [sys.executable, str(script_path)],
                        capture_output=True,
                        text=True,
                        encoding="utf-8",
                        errors="replace",
                        cwd=td,
                        env=env,
                        timeout=180,
                    )
                except Exception as e:
                    last_error = f"执行爬虫失败：{e}"
                    error_for_prompt = last_error
                else:
                    stdout = (proc.stdout or "").strip()
                    stderr = (proc.stderr or "").strip()
                    if proc.returncode == 0 and output_path.exists():
                        try:
                            df = pd.read_csv(output_path)
                        except Exception as e:
                            last_error = f"读取爬取结果失败：{e}"
                            error_for_prompt = last_error
                        else:
                            self.data_ready.emit(df, self.url)
                            return
                    else:
                        if proc.returncode != 0:
                            last_error = stderr or stdout or "未知错误"
                        elif not output_path.exists():
                            last_error = "脚本执行完成但未生成CSV文件"
                        else:
                            last_error = stderr or stdout or "未知错误"
                        error_for_prompt = stderr or stdout or last_error

                if attempt == max_attempts - 1:
                    break

                condensed_error = summarize_error(error_for_prompt)
                prompt_error = error_for_prompt if len(error_for_prompt) <= 1500 else error_for_prompt[:1500] + "…"
                messages.append({"role": "assistant", "content": f"```python\n{code}\n```"})
                messages.append({
                    "role": "user",
                    "content": (
                        "脚本执行失败。错误摘要："
                        f"{condensed_error}\n详细日志：{prompt_error}\n"
                        "请基于上述代码修复问题，并继续遵守所有约束，直到成功导出CSV文件。"
                    ),
                })

        self.error.emit(f"爬虫生成失败：{last_error or '未知错误'}")


class ScraperDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("从网页获取数据")
        layout = QVBoxLayout(self)
        notice = QLabel(
            "提示：当前功能主要适用于结构稳定的静态或半静态网页。"
            "若页面依赖复杂的JS交互、登录或反爬机制，脚本可能失败；未来将考虑接入 Selenium/Playwright 等浏览器自动化方案。"
        )
        notice.setWordWrap(True)
        notice.setStyleSheet("color: #94a3b8;")
        layout.addWidget(notice)
        layout.addWidget(QLabel("目标网址:"))
        self.url_edit = QLineEdit()
        layout.addWidget(self.url_edit)
        layout.addWidget(QLabel("爬取需求描述:"))
        self.instruction_edit = QPlainTextEdit()
        self.instruction_edit.setPlaceholderText("例如：抓取此页面的商品列表与价格，并保存为csv。")
        self.instruction_edit.setMinimumHeight(120)
        layout.addWidget(self.instruction_edit)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._on_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _on_accept(self):
        url = self.url_edit.text().strip()
        if not url:
            QMessageBox.warning(self, "提示", "请输入有效的网址")
            return
        instruction = self.instruction_edit.toPlainText().strip()
        if not instruction:
            QMessageBox.warning(self, "提示", "请输入爬取需求")
            return
        self.accept()

    def get_values(self) -> tuple[str, str]:
        return self.url_edit.text().strip(), self.instruction_edit.toPlainText().strip()


class DataReporterWidget(QWidget):
    def __init__(self, settings: AppSettings, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.settings = settings
        self.current_df: Optional[pd.DataFrame] = None
        self.sanitized_df: Optional[pd.DataFrame] = None
        self.current_source: str = ""
        self.current_path: Optional[Path] = None
        self.report_worker: Optional[ReportGenerationWorker] = None
        self.scraper_worker: Optional[ScraperWorker] = None
        self.quick_worker: Optional[IntentWorker] = None
        self.quick_buttons: List[QPushButton] = []
        self.quick_placeholder: Optional[QLabel] = None
        self._quick_request_serial: int = 0
        self._latest_detection_hint: str = ""
        self.latest_report_html: str = ""
        self.latest_report_code: str = ""
        self.setAcceptDrops(True)
        self._build_ui()

    def _build_ui(self):
        main_layout = QHBoxLayout(self)
        main_layout.setContentsMargins(12, 12, 12, 12)
        main_layout.setSpacing(12)

        left_container = QWidget()
        left_layout = QVBoxLayout(left_container)
        left_layout.setSpacing(10)

        source_group = QGroupBox("数据源")
        source_layout = QVBoxLayout(source_group)
        btn_row = QHBoxLayout()
        self.btn_select_file = QPushButton("选择本地文件…")
        self.btn_select_file.clicked.connect(self.select_file)
        btn_row.addWidget(self.btn_select_file)
        self.btn_scrape = QPushButton("从网页获取数据")
        self.btn_scrape.clicked.connect(self.fetch_from_web)
        btn_row.addWidget(self.btn_scrape)
        source_layout.addLayout(btn_row)
        self.source_label = QLabel("未选择数据源")
        self.source_label.setWordWrap(True)
        source_layout.addWidget(self.source_label)
        self.summary_label = QLabel("")
        source_layout.addWidget(self.summary_label)
        left_layout.addWidget(source_group)

        preview_group = QGroupBox("数据预览")
        preview_layout = QVBoxLayout(preview_group)
        self.preview_table = QTableWidget()
        self.preview_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.preview_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.preview_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.preview_table.horizontalHeader().setStretchLastSection(True)
        preview_layout.addWidget(self.preview_table)
        data_btn_row = QHBoxLayout()
        self.btn_export_data = QPushButton("导出当前数据…")
        self.btn_export_data.setEnabled(False)
        self.btn_export_data.clicked.connect(self._export_current_data)
        data_btn_row.addWidget(self.btn_export_data)
        data_btn_row.addStretch()
        preview_layout.addLayout(data_btn_row)
        left_layout.addWidget(preview_group, 1)

        instruction_group = QGroupBox("分析需求")
        instruction_layout = QVBoxLayout(instruction_group)
        quick_row = QHBoxLayout()
        quick_label = QLabel("快速指令：")
        quick_label.setStyleSheet("color: #94a3b8;")
        quick_row.addWidget(quick_label)
        self.quick_container = QWidget()
        self.quick_container_layout = QHBoxLayout(self.quick_container)
        self.quick_container_layout.setContentsMargins(0, 0, 0, 0)
        self.quick_container_layout.setSpacing(6)
        quick_row.addWidget(self.quick_container, 1)
        instruction_layout.addLayout(quick_row)
        self._set_quick_placeholder("加载数据后将自动推荐常见分析任务。")
        self.instruction_edit = QPlainTextEdit()
        self.instruction_edit.setPlaceholderText("例如：分析各产品销售额占比，并给出关键洞察。")
        self.instruction_edit.setMinimumHeight(120)
        instruction_layout.addWidget(self.instruction_edit)
        self.btn_generate = QPushButton("生成报告")
        self.btn_generate.clicked.connect(self.generate_report)
        instruction_layout.addWidget(self.btn_generate)
        self.status_label = QLabel("请先选择数据源")
        instruction_layout.addWidget(self.status_label)
        left_layout.addWidget(instruction_group)

        left_layout.addStretch()

        right_container = QWidget()
        right_layout = QVBoxLayout(right_container)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(6)
        report_group = QGroupBox("报告与代码")
        report_layout = QVBoxLayout(report_group)
        self.report_tabs = QTabWidget()
        self.report_view = QWebEngineView()
        self.report_view.setHtml("<h3 style='color:#94a3b8;'>生成的报告将在此显示</h3>")
        self.report_tabs.addTab(self.report_view, "报表预览")
        self.code_view = QPlainTextEdit()
        self.code_view.setReadOnly(True)
        self.code_view.setPlaceholderText("生成的 Python 代码将显示在此处。")
        self.report_tabs.addTab(self.code_view, "生成代码")
        report_layout.addWidget(self.report_tabs)
        report_btn_row = QHBoxLayout()
        self.btn_export_report_html = QPushButton("导出报表…")
        self.btn_export_report_html.setEnabled(False)
        self.btn_export_report_html.clicked.connect(self._export_report_html)
        report_btn_row.addWidget(self.btn_export_report_html)
        self.btn_export_report_code = QPushButton("导出代码…")
        self.btn_export_report_code.setEnabled(False)
        self.btn_export_report_code.clicked.connect(self._export_report_code)
        report_btn_row.addWidget(self.btn_export_report_code)
        report_btn_row.addStretch()
        report_layout.addLayout(report_btn_row)
        right_layout.addWidget(report_group, 1)

        main_layout.addWidget(left_container, 1)
        main_layout.addWidget(right_container, 1)

    # ---- 数据加载与预览 ----
    def select_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择数据文件",
            "",
            "数据文件 (*.xlsx *.xlsm *.xls *.csv)"
        )
        if path:
            self._load_file(Path(path))

    def fetch_from_web(self):
        if self.scraper_worker and self.scraper_worker.isRunning():
            QMessageBox.information(self, "请稍候", "当前正在执行爬虫任务，请等待完成。")
            return
        dlg = ScraperDialog(self)
        if dlg.exec():
            url, instruction = dlg.get_values()
            api_key = (self.settings.get("report_api_key", "") or "").strip()
            worker = ScraperWorker(api_key, url, instruction)
            worker.progress.connect(self._set_status)
            worker.error.connect(self._handle_scraper_error)
            worker.data_ready.connect(self._handle_scraper_success)
            worker.finished.connect(self._scraper_finished)
            self.scraper_worker = worker
            self._set_status("正在启动智能爬虫…")
            worker.start()

    def _scraper_finished(self):
        self.scraper_worker = None

    def _handle_scraper_error(self, message: str):
        self.scraper_worker = None
        QMessageBox.critical(self, "爬虫错误", message)
        self._set_status("爬虫执行失败")

    def _handle_scraper_success(self, df: pd.DataFrame, url: str):
        self.scraper_worker = None
        self._adopt_dataframe(df, f"来自网页数据：{url}")
        self._set_status_with_hint("已加载爬取数据，请输入分析需求。")

    def _load_file(self, path: Path):
        detection_hint = ""
        if path.suffix.lower() == ".csv":
            try:
                df = pd.read_csv(path)
            except Exception as e:
                QMessageBox.critical(self, "错误", f"读取文件失败：{e}")
                return
            detection_hint = "CSV 文件默认使用首行作为表头。"
        else:
            try:
                sheet_names = get_sheet_names_safely(path)
            except Exception as e:
                QMessageBox.critical(self, "错误", f"读取工作表失败：{e}")
                return
            if not sheet_names:
                QMessageBox.critical(self, "错误", "该工作簿中没有可用的工作表。")
                return
            sheet_name = sheet_names[0]
            try:
                header_row, start_col = auto_detect_header_start(path, sheet_name)
                detection_hint = f"自动识别表头：第{header_row}行，数据起始列：第{start_col}列。"
                structure_df = read_excel_dataframe(path, sheet_name, header_row, start_col)
                df = pd.read_excel(
                    path,
                    sheet_name=sheet_name,
                    header=header_row - 1,
                    engine="openpyxl",
                )
                if start_col > 1:
                    df = df.iloc[:, start_col - 1 :]
                df = df.dropna(how="all")
                df.columns = structure_df.columns
            except Exception as detect_err:
                try:
                    df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
                except Exception as e:
                    QMessageBox.critical(self, "错误", f"读取文件失败：{e}")
                    return
                detail = str(detect_err).strip().splitlines()[0]
                if detail:
                    detection_hint = f"自动识别表头失败，已按默认方式读取。原因：{detail}"
                else:
                    detection_hint = "自动识别表头失败，已按默认方式读取。"

        self._adopt_dataframe(df, f"本地文件：{path.name}", path, detection_hint)
        self._set_status_with_hint("已加载本地数据，请输入分析需求。")

    def _adopt_dataframe(self, df: pd.DataFrame, source_desc: str, source_path: Optional[Path] = None, detection_hint: str = ""):
        self.current_df = df.copy()
        self.sanitized_df, _ = desensitize_dataframe(df)
        self.current_source = source_desc
        self.current_path = source_path
        self._latest_detection_hint = detection_hint.strip()
        self.source_label.setText(source_desc)
        rows, cols = df.shape
        self.summary_label.setText(f"记录数：{rows} 行 × {cols} 列")
        self._populate_preview(df)
        self.report_view.setHtml("<h3 style='color:#94a3b8;'>请生成新的报告</h3>")
        if hasattr(self, "code_view"):
            self.code_view.clear()
        self.latest_report_html = ""
        self.latest_report_code = ""
        if hasattr(self, "btn_export_report_html"):
            self.btn_export_report_html.setEnabled(False)
        if hasattr(self, "btn_export_report_code"):
            self.btn_export_report_code.setEnabled(False)
        if hasattr(self, "btn_export_data"):
            self.btn_export_data.setEnabled(True)
        self._trigger_quick_suggestions()

    def _populate_preview(self, df: pd.DataFrame):
        if df is None:
            self.preview_table.clear()
            self.preview_table.setRowCount(0)
            self.preview_table.setColumnCount(0)
            return

        preview = df.head(200)
        self.preview_table.setRowCount(len(preview.index))
        self.preview_table.setColumnCount(len(preview.columns))
        self.preview_table.setHorizontalHeaderLabels([str(c) for c in preview.columns])
        for r, (_, row) in enumerate(preview.iterrows()):
            for c, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                self.preview_table.setItem(r, c, item)
        self.preview_table.resizeColumnsToContents()

    def _clear_quick_container(self):
        if not hasattr(self, "quick_container_layout"):
            return
        while self.quick_container_layout.count():
            item = self.quick_container_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

    def _set_quick_placeholder(self, text: str):
        self.quick_buttons.clear()
        self._clear_quick_container()
        label = QLabel(text)
        label.setStyleSheet("color: #94a3b8;")
        self.quick_placeholder = label
        self.quick_container_layout.addWidget(label)
        self.quick_container_layout.addStretch()

    def _show_quick_buttons(self, texts: List[str]):
        self._clear_quick_container()
        self.quick_buttons.clear()
        for text in texts:
            btn = QPushButton(text)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.clicked.connect(lambda _, t=text: self._apply_preset_instruction(t))
            self.quick_container_layout.addWidget(btn)
            self.quick_buttons.append(btn)
        self.quick_placeholder = None
        self.quick_container_layout.addStretch()

    def _stop_quick_worker(self):
        if self.quick_worker and self.quick_worker.isRunning():
            self.quick_worker.requestInterruption()
            self.quick_worker.wait(200)
        self.quick_worker = None

    def _trigger_quick_suggestions(self):
        if not hasattr(self, "quick_container_layout"):
            return
        self._stop_quick_worker()
        if self.current_df is None:
            self._set_quick_placeholder("请先选择数据源。")
            return
        if self.current_df.empty:
            self._set_quick_placeholder("当前数据为空，无法推荐分析指令。")
            return
        api_key = (self.settings.get("report_api_key", "") or "").strip()
        if not api_key:
            self._set_quick_placeholder("请在设置中配置数据报表模块的 API Key。")
            return

        self._quick_request_serial += 1
        request_id = self._quick_request_serial
        entry = {
            "dataframe": self.current_df.head(200).copy(),
            "name": self.current_source or (self.current_path.name if self.current_path else "数据集"),
        }
        if self.current_path:
            entry["path"] = str(self.current_path)
        self._set_quick_placeholder("正在分析常见分析需求…")
        worker = IntentWorker(api_key, [entry])
        worker.results.connect(lambda intents, rid=request_id: self._on_quick_intents_ready(rid, intents))
        worker.error.connect(lambda message, rid=request_id: self._on_quick_intents_error(rid, message))
        worker.finished.connect(self._on_quick_worker_finished)
        self.quick_worker = worker
        worker.start()

    def _on_quick_intents_ready(self, request_id: int, intents: List[str]):
        if request_id != self._quick_request_serial:
            return
        cleaned = []
        for item in intents:
            text = str(item).strip()
            if text:
                cleaned.append(text)
        if not cleaned:
            self._set_quick_placeholder("未能识别常见分析指令，请手动描述需求。")
            return
        self._show_quick_buttons(cleaned[:5])
        self._set_status_with_hint("已推荐常见分析需求，可直接点击或继续编辑。")

    def _on_quick_intents_error(self, request_id: int, message: str):
        if request_id != self._quick_request_serial:
            return
        self._set_quick_placeholder("未能获取推荐指令，请手动描述需求。")
        first_line = (message or "").splitlines()[0] if message else ""
        if first_line:
            self._set_status(f"推荐指令获取失败：{first_line}")
        else:
            self._set_status("推荐指令获取失败。")

    def _on_quick_worker_finished(self):
        self.quick_worker = None

    def _set_status_with_hint(self, base_text: str):
        base = base_text.strip()
        hint = self._latest_detection_hint.strip()
        if hint:
            if base:
                self._set_status(f"{base} {hint}")
            else:
                self._set_status(hint)
        else:
            self._set_status(base)

    # ---- 报告生成 ----
    def generate_report(self):
        if self.report_worker and self.report_worker.isRunning():
            QMessageBox.information(self, "请稍候", "当前已有生成任务在执行，请稍候完成。")
            return
        if self.current_df is None or self.sanitized_df is None:
            QMessageBox.warning(self, "提示", "请先选择或获取数据。")
            return

        instruction = self.instruction_edit.toPlainText().strip()
        if not instruction:
            QMessageBox.warning(self, "提示", "请输入报告需求。")
            return

        api_key = (self.settings.get("report_api_key", "") or "").strip()
        if not api_key:
            QMessageBox.warning(self, "提示", "请先在设置中配置数据报表模块的 API Key。")
            return

        sample_csv = self.sanitized_df.head(20).to_csv(index=False)
        worker = ReportGenerationWorker(api_key, sample_csv, list(self.current_df.columns), instruction, self.current_df.copy())
        worker.progress.connect(self._set_status)
        worker.error.connect(self._handle_report_error)
        worker.code_ready.connect(self._handle_report_code)
        worker.success.connect(self._handle_report_success)
        worker.finished.connect(self._report_finished)
        self.report_worker = worker
        self.btn_generate.setEnabled(False)
        if hasattr(self, "code_view"):
            self.code_view.setPlainText("正在调用模型生成报告代码…")
            if hasattr(self, "report_tabs"):
                self.report_tabs.setCurrentWidget(self.code_view)
        self._set_status("正在生成数据报告…")
        worker.start()

    def _handle_report_code(self, code: str):
        if not hasattr(self, "code_view"):
            return
        display = code or ""
        self.code_view.setPlainText(display)
        self.latest_report_code = code
        if hasattr(self, "btn_export_report_code"):
            self.btn_export_report_code.setEnabled(bool(code.strip()))
        sb = self.code_view.verticalScrollBar()
        if sb is not None:
            sb.setValue(sb.maximum())

    def _handle_report_success(self, html: str, code: str):
        self._handle_report_code(code)
        self.report_view.setHtml(html)
        self.latest_report_html = html
        if hasattr(self, "btn_export_report_html"):
            self.btn_export_report_html.setEnabled(bool((html or "").strip()))
        if hasattr(self, "report_tabs"):
            self.report_tabs.setCurrentWidget(self.report_view)
        self._set_status("报告生成完成")

    def _handle_report_error(self, message: str):
        QMessageBox.critical(self, "生成失败", message)
        self._set_status("报告生成失败")

    def _report_finished(self):
        self.btn_generate.setEnabled(True)
        self.report_worker = None

    def _export_report_html(self):
        html = (self.latest_report_html or "").strip()
        if not html:
            QMessageBox.information(self, "提示", "当前没有可导出的报表，请先生成报告。")
            return
        default_path = self._default_export_path("数据分析报告.html")
        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出数据报表",
            default_path,
            "HTML 文件 (*.html *.htm)",
        )
        if not path:
            return
        export_path = Path(path)
        if export_path.suffix.lower() not in {".html", ".htm"}:
            export_path = export_path.with_suffix(".html")
        try:
            export_path.write_text(html, encoding="utf-8")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"写入文件失败：{e}")
            return
        self.settings.update(last_ai_export_path=str(export_path.parent))
        self._set_status(f"报表已导出：{export_path}")
        QMessageBox.information(self, "导出完成", f"报表已导出至：\n{export_path}")

    def _export_report_code(self):
        code = (self.latest_report_code or "").strip()
        if not code:
            QMessageBox.information(self, "提示", "当前没有可导出的代码，请先生成报告。")
            return
        default_path = self._default_export_path("report_generator.py")
        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出生成代码",
            default_path,
            "Python 文件 (*.py)",
        )
        if not path:
            return
        export_path = Path(path)
        if export_path.suffix.lower() != ".py":
            export_path = export_path.with_suffix(".py")
        try:
            export_path.write_text(self.latest_report_code, encoding="utf-8")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"写入文件失败：{e}")
            return
        self.settings.update(last_ai_export_path=str(export_path.parent))
        self._set_status(f"代码已导出：{export_path}")
        QMessageBox.information(self, "导出完成", f"代码已导出至：\n{export_path}")

    def _export_current_data(self):
        if self.current_df is None:
            QMessageBox.information(self, "提示", "当前没有可导出的数据集。")
            return
        default_filename = "导出数据.csv"
        default_path = self._default_export_path(default_filename)
        filters = "CSV 文件 (*.csv);;Excel 工作簿 (*.xlsx)"
        path, selected_filter = QFileDialog.getSaveFileName(
            self,
            "导出当前数据",
            default_path,
            filters,
        )
        if not path:
            return
        export_path = Path(path)
        chosen_filter = (selected_filter or "").lower()
        try:
            if "xlsx" in chosen_filter:
                if export_path.suffix.lower() != ".xlsx":
                    export_path = export_path.with_suffix(".xlsx")
                self.current_df.to_excel(export_path, index=False)
            else:
                if export_path.suffix.lower() != ".csv":
                    export_path = export_path.with_suffix(".csv")
                self.current_df.to_csv(export_path, index=False, encoding="utf-8-sig")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"保存数据失败：{e}")
            return
        self.settings.update(last_ai_export_path=str(export_path.parent))
        self._set_status(f"数据已导出：{export_path}")
        QMessageBox.information(self, "导出完成", f"数据已导出至：\n{export_path}")

    def _default_export_path(self, filename: str) -> str:
        base_dir = (self.settings.get("last_ai_export_path", "") or "").strip()
        if base_dir:
            try:
                base_path = Path(base_dir)
                if base_path.exists() and not base_path.is_dir():
                    base_path = base_path.parent
                if base_path.exists() or base_path.parent.exists():
                    return str(base_path / filename)
            except Exception:
                pass
        return filename

    def _apply_preset_instruction(self, text: str):
        self.instruction_edit.setPlainText(text)
        self._set_status_with_hint("已应用快速指令，可直接生成或继续编辑。")

    def _set_status(self, text: str):
        self.status_label.setText(text)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            super().dragEnterEvent(event)

    def dropEvent(self, event):
        for url in event.mimeData().urls():
            local_path = url.toLocalFile()
            if local_path:
                self._load_file(Path(local_path))
                break
        event.acceptProposedAction()
        super().dropEvent(event)


class DataCleanerWidget(QWidget):
    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.current_df: Optional[pd.DataFrame] = None
        self.current_path: Optional[Path] = None
        self._build_ui()

    def _build_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(12, 12, 12, 12)
        main_layout.setSpacing(12)

        top_row = QHBoxLayout()
        self.btn_open = QPushButton("加载数据…")
        self.btn_open.clicked.connect(self.open_file)
        top_row.addWidget(self.btn_open)
        self.btn_export = QPushButton("导出清洗结果")
        self.btn_export.clicked.connect(self.export_file)
        top_row.addWidget(self.btn_export)
        top_row.addStretch()
        main_layout.addLayout(top_row)

        splitter = QSplitter()
        splitter.setOrientation(Qt.Orientation.Horizontal)

        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(8)
        self.path_label = QLabel("未加载数据")
        self.path_label.setWordWrap(True)
        left_layout.addWidget(self.path_label)
        self.clean_preview = QTableWidget()
        self.clean_preview.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.clean_preview.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.clean_preview.horizontalHeader().setStretchLastSection(True)
        left_layout.addWidget(self.clean_preview, 1)
        self.status_label = QLabel("请选择数据并执行清洗操作。")
        left_layout.addWidget(self.status_label)

        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(10)

        self._build_dedup_section(right_layout)
        self._build_null_section(right_layout)
        self._build_text_section(right_layout)
        self._build_convert_section(right_layout)
        right_layout.addStretch()

        splitter.addWidget(left_widget)
        splitter.addWidget(right_widget)
        splitter.setSizes([700, 400])
        main_layout.addWidget(splitter, 1)

    def _build_dedup_section(self, layout: QVBoxLayout):
        group = QGroupBox("去重")
        group_layout = QFormLayout(group)
        self.dedup_cols_edit = QLineEdit()
        self.dedup_cols_edit.setPlaceholderText("留空表示按所有列去重；多个列名用逗号分隔")
        group_layout.addRow("列名称:", self.dedup_cols_edit)
        btn = QPushButton("执行去重")
        btn.clicked.connect(self.apply_dedup)
        group_layout.addRow(btn)
        layout.addWidget(group)

    def _build_null_section(self, layout: QVBoxLayout):
        group = QGroupBox("空值处理")
        form = QFormLayout(group)
        self.dropna_cols_edit = QLineEdit()
        self.dropna_cols_edit.setPlaceholderText("留空表示删除任何包含空值的行")
        btn_drop = QPushButton("删除空值行")
        btn_drop.clicked.connect(self.apply_dropna)
        form.addRow("列名称:", self.dropna_cols_edit)
        form.addRow(btn_drop)

        self.fillna_cols_edit = QLineEdit()
        self.fillna_cols_edit.setPlaceholderText("多个列名用逗号分隔，留空表示全部列")
        form.addRow("填充列:", self.fillna_cols_edit)
        self.fillna_method_combo = QComboBox()
        self.fillna_method_combo.addItem("使用指定文本", "custom")
        self.fillna_method_combo.addItem("填充为0", "zero")
        self.fillna_method_combo.addItem("填充列平均值", "mean")
        self.fillna_method_combo.addItem("填充列中位数", "median")
        self.fillna_method_combo.addItem("填充列众数", "mode")
        form.addRow("填充方式:", self.fillna_method_combo)
        self.fillna_value_edit = QLineEdit()
        self.fillna_value_edit.setPlaceholderText("当选择指定文本时填写此处")
        form.addRow("填充值:", self.fillna_value_edit)
        btn_fill = QPushButton("填充空值")
        btn_fill.clicked.connect(self.apply_fillna)
        form.addRow(btn_fill)
        layout.addWidget(group)

    def _build_text_section(self, layout: QVBoxLayout):
        group = QGroupBox("文本处理")
        form = QFormLayout(group)
        self.trim_cols_edit = QLineEdit()
        self.trim_cols_edit.setPlaceholderText("多个列名用逗号分隔，留空表示所有字符串列")
        form.addRow("去除首尾空格:", self.trim_cols_edit)
        btn = QPushButton("执行修剪")
        btn.clicked.connect(self.apply_trim)
        form.addRow(btn)
        layout.addWidget(group)

    def _build_convert_section(self, layout: QVBoxLayout):
        group = QGroupBox("格式转换")
        form = QFormLayout(group)
        self.convert_col_edit = QLineEdit()
        form.addRow("目标列:", self.convert_col_edit)
        self.convert_type_combo = QComboBox()
        self.convert_type_combo.addItem("文本", "str")
        self.convert_type_combo.addItem("整数", "int")
        self.convert_type_combo.addItem("浮点数", "float")
        form.addRow("目标类型:", self.convert_type_combo)
        btn = QPushButton("转换类型")
        btn.clicked.connect(self.apply_convert)
        form.addRow(btn)
        layout.addWidget(group)

    # ---- 数据加载与导出 ----
    def open_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择数据文件",
            "",
            "数据文件 (*.xlsx *.xlsm *.xls *.csv)"
        )
        if path:
            detection_hint = ""
            try:
                if path.lower().endswith(".csv"):
                    df = pd.read_csv(path)
                    detection_hint = "CSV 文件默认使用首行作为表头。"
                else:
                    path_obj = Path(path)
                    try:
                        sheet_names = get_sheet_names_safely(path_obj)
                    except Exception as e:
                        QMessageBox.critical(self, "错误", f"读取工作表失败：{e}")
                        return
                    if not sheet_names:
                        QMessageBox.critical(self, "错误", "该工作簿中没有可用的工作表。")
                        return
                    sheet_name = sheet_names[0]
                    try:
                        header_row, start_col = auto_detect_header_start(path_obj, sheet_name)
                        detection_hint = f"自动识别表头：第{header_row}行，数据起始列：第{start_col}列。"
                        structure_df = read_excel_dataframe(path_obj, sheet_name, header_row, start_col)
                        df = pd.read_excel(
                            path_obj,
                            sheet_name=sheet_name,
                            header=header_row - 1,
                            engine="openpyxl",
                        )
                        if start_col > 1:
                            df = df.iloc[:, start_col - 1 :]
                        df = df.dropna(how="all")
                        df.columns = structure_df.columns
                    except Exception as detect_err:
                        try:
                            df = pd.read_excel(path_obj, sheet_name=sheet_name, engine="openpyxl")
                        except Exception as e:
                            QMessageBox.critical(self, "错误", f"读取文件失败：{e}")
                            return
                        detail = str(detect_err).strip().splitlines()[0]
                        if detail:
                            detection_hint = f"自动识别表头失败，已按默认方式读取。原因：{detail}"
                        else:
                            detection_hint = "自动识别表头失败，已按默认方式读取。"
            except Exception as e:
                QMessageBox.critical(self, "错误", f"读取文件失败：{e}")
                return
            self.current_df = df
            self.current_path = Path(path)
            self.path_label.setText(f"当前文件：{Path(path).name}")
            self._refresh_preview()
            status = "数据已加载，可选择清洗操作。"
            if detection_hint:
                status = f"{status} {detection_hint}"
            self._set_status(status.strip())

    def export_file(self):
        if self.current_df is None:
            QMessageBox.warning(self, "提示", "请先加载并处理数据。")
            return
        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出结果",
            "",
            "Excel 文件 (*.xlsx)"
        )
        if not path:
            return
        try:
            self.current_df.to_excel(path, index=False)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败：{e}")
            return
        self._set_status(f"已导出至：{path}")

    # ---- 操作实现 ----
    def _parse_columns(self, text: str) -> Optional[List[str]]:
        cols = [c.strip() for c in text.split(",") if c.strip()]
        return cols if cols else None

    def _ensure_dataframe(self) -> bool:
        if self.current_df is None:
            QMessageBox.warning(self, "提示", "请先加载数据。")
            return False
        return True

    def apply_dedup(self):
        if not self._ensure_dataframe():
            return
        cols = self._parse_columns(self.dedup_cols_edit.text())
        try:
            if cols:
                missing = [c for c in cols if c not in self.current_df.columns]
                if missing:
                    QMessageBox.warning(self, "提示", f"以下列不存在：{missing}")
                    return
                self.current_df = self.current_df.drop_duplicates(subset=cols)
            else:
                self.current_df = self.current_df.drop_duplicates()
            self._refresh_preview()
            self._set_status("已完成去重处理。")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"去重失败：{e}")

    def apply_dropna(self):
        if not self._ensure_dataframe():
            return
        cols = self._parse_columns(self.dropna_cols_edit.text())
        try:
            if cols:
                missing = [c for c in cols if c not in self.current_df.columns]
                if missing:
                    QMessageBox.warning(self, "提示", f"以下列不存在：{missing}")
                    return
                self.current_df = self.current_df.dropna(subset=cols)
            else:
                self.current_df = self.current_df.dropna()
            self._refresh_preview()
            self._set_status("已删除包含空值的行。")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"删除空值失败：{e}")

    def apply_fillna(self):
        if not self._ensure_dataframe():
            return
        cols = self._parse_columns(self.fillna_cols_edit.text())
        target_cols = cols if cols else list(self.current_df.columns)
        missing = [c for c in target_cols if c not in self.current_df.columns]
        if missing:
            QMessageBox.warning(self, "提示", f"以下列不存在：{missing}")
            return

        method = self.fillna_method_combo.currentData()
        try:
            if method == "custom":
                value = self.fillna_value_edit.text()
                self.current_df[target_cols] = self.current_df[target_cols].fillna(value)
            elif method == "zero":
                self.current_df[target_cols] = self.current_df[target_cols].fillna(0)
            else:
                for col in target_cols:
                    series = self.current_df[col]
                    numeric = pd.to_numeric(series, errors="coerce")
                    if method == "mean":
                        fill_value = numeric.mean()
                    elif method == "median":
                        fill_value = numeric.median()
                    else:
                        fill_value = series.mode().iloc[0] if not series.mode().empty else ""
                    self.current_df[col] = series.fillna(fill_value)
            self._refresh_preview()
            self._set_status("已完成空值填充。")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"填充失败：{e}")

    def apply_trim(self):
        if not self._ensure_dataframe():
            return
        cols = self._parse_columns(self.trim_cols_edit.text())
        target_cols = cols if cols else [c for c in self.current_df.columns if self.current_df[c].dtype == object]
        missing = [c for c in target_cols if c not in self.current_df.columns]
        if missing:
            QMessageBox.warning(self, "提示", f"以下列不存在：{missing}")
            return
        try:
            for col in target_cols:
                self.current_df[col] = self.current_df[col].astype(str).str.strip()
            self._refresh_preview()
            self._set_status("已完成文本修剪。")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"文本修剪失败：{e}")

    def apply_convert(self):
        if not self._ensure_dataframe():
            return
        col = self.convert_col_edit.text().strip()
        if not col:
            QMessageBox.warning(self, "提示", "请填写需要转换的列名。")
            return
        if col not in self.current_df.columns:
            QMessageBox.warning(self, "提示", f"列 {col} 不存在。")
            return
        target_type = self.convert_type_combo.currentData()
        try:
            if target_type == "str":
                self.current_df[col] = self.current_df[col].astype(str)
            elif target_type == "int":
                self.current_df[col] = pd.to_numeric(self.current_df[col], errors="coerce").astype("Int64")
            else:
                self.current_df[col] = pd.to_numeric(self.current_df[col], errors="coerce").astype(float)
            self._refresh_preview()
            self._set_status("已完成格式转换。")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"转换失败：{e}")

    # ---- UI 辅助 ----
    def _refresh_preview(self):
        if self.current_df is None:
            self.clean_preview.clear()
            self.clean_preview.setRowCount(0)
            self.clean_preview.setColumnCount(0)
            return
        preview = self.current_df.head(200)
        self.clean_preview.setRowCount(len(preview.index))
        self.clean_preview.setColumnCount(len(preview.columns))
        self.clean_preview.setHorizontalHeaderLabels([str(c) for c in preview.columns])
        for r, (_, row) in enumerate(preview.iterrows()):
            for c, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                self.clean_preview.setItem(r, c, item)
        self.clean_preview.resizeColumnsToContents()

    def _set_status(self, text: str):
        self.status_label.setText(text)


# ===================== 主窗口 =====================


class SmartVlookuperMainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.settings = AppSettings()
        self.setWindowTitle("诗忻智能Excel数据伴侣 v2.0")
        self.resize(1400, 900)
        self._build_ui()
        self._create_menus()
        self.apply_theme()

    def _build_ui(self):
        central = QWidget()
        layout = QHBoxLayout(central)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self.nav_list = QListWidget()
        self.nav_list.setFixedWidth(200)
        self.nav_list.setSpacing(2)
        self.nav_list.addItem(QListWidgetItem("数据匹配"))
        self.nav_list.addItem(QListWidgetItem("数据报表"))
        self.nav_list.addItem(QListWidgetItem("数据清洗"))
        self.nav_list.addItem(QListWidgetItem("AI 助手"))
        layout.addWidget(self.nav_list)

        self.stack = QStackedWidget()
        self.data_matching = DataMatchingWidget(self.settings, self)
        self.data_reporter = DataReporterWidget(self.settings, self)
        self.data_cleaner = DataCleanerWidget(self)
        self.ai_widget = AIAssistantWidget(self.settings, self)
        self.stack.addWidget(self.data_matching)
        self.stack.addWidget(self.data_reporter)
        self.stack.addWidget(self.data_cleaner)
        self.stack.addWidget(self.ai_widget)
        layout.addWidget(self.stack, 1)

        self.nav_list.currentRowChanged.connect(self.stack.setCurrentIndex)
        self.nav_list.setCurrentRow(0)

        self.setCentralWidget(central)

    def _create_menus(self):
        menu_settings = self.menuBar().addMenu("设置")
        action_prefs = QAction("首选项…", self)
        action_prefs.triggered.connect(self.open_settings_dialog)
        menu_settings.addAction(action_prefs)

        menu_about = self.menuBar().addMenu("关于")
        action_about = QAction("关于本应用", self)
        action_about.triggered.connect(self.show_about_dialog)
        menu_about.addAction(action_about)

    def apply_theme(self):
        theme = self.settings.get("theme", "dark")
        style = DARK_STYLESHEET if theme == "dark" else LIGHT_STYLESHEET
        app = QApplication.instance()
        if app:
            app.setStyleSheet(style)

    def open_settings_dialog(self):
        dlg = SettingsDialog(self, self.settings)
        if dlg.exec():
            self.apply_theme()

    def show_about_dialog(self):
        QMessageBox.information(
            self,
            "关于",
            "诗忻智能Excel数据伴侣 v2.0\n© 梁诗忻 2025. 本项目采用 MIT 许可证。\n项目地址：https://github.com/liangshixin1/Smart-Vlookuper"
        )


def main():
    app = QApplication(sys.argv)
    w = SmartVlookuperMainWindow()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
